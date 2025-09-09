import base64
import re
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse, JsonResponse
from django.db import transaction, connection
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.template.loader import render_to_string
import json
import pytz
import bcrypt
import os
from django.conf import settings as django_settings
import logging  
from io import BytesIO
from urllib.parse import urljoin
from weasyprint import HTML, default_url_fetcher
from django.http import JsonResponse, HttpResponse
from decimal import Decimal
from datetime import datetime, timedelta
from django.contrib.auth.hashers import check_password
from django.shortcuts import render
from django.http import Http404
from django.contrib import messages
from collections import namedtuple
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User 
import os
import random
import string
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from django.conf import settings
from django.db import connection, transaction, DatabaseError

from django.db import connection
from django.contrib.auth.models import AnonymousUser
from django.shortcuts import render
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from datetime import datetime

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.db import connection
from datetime import datetime
import pytz
import json
import traceback
# Set up logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

def base(request):
    """Render the base page"""
    return render(request, 'base.html')

def abipet(request):
    return render(request, 'pettemplate.html')

def index(request):
    """Render the login page"""
    return render(request, 'login.html')

def report(request):
    """Render the report page"""
    return render(request, 'report.html')

@ensure_csrf_cookie
def new_username(request):
    """Check if Employee ID exists and fetch user details"""
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id', '').strip()  

        if not employee_id:
            return JsonResponse({'status': 'error', 'message': 'Employee ID is required'})

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT id, employee_id, username, role FROM users WHERE employee_id = %s", 
                [employee_id]
            )
            result = cursor.fetchone()

        if result:
            user_id, employee_id, username, role = result
            request.session['user_id'] = user_id
            request.session['employee_id'] = employee_id
            request.session['username'] = username
            request.session['role'] = role

            return JsonResponse({
                'status': 'success',
                'user_id': user_id,
                'employee_id': employee_id,
                'username': username,
                'role': role
            })
        else:
            return JsonResponse({'status': 'error', 'message': 'Employee ID not found'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})



@ensure_csrf_cookie
def new_pwd(request):
    """Authenticate user and perform Django login"""
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id', '').strip()
        password = request.POST.get('password', '').strip()

        if not employee_id or not password:
            return JsonResponse({'status': 'error', 'message': 'Missing EmployeeId or password'})

        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT id, employee_id, username, password, role FROM users WHERE employee_id = %s", [employee_id])
                result = cursor.fetchone()

            if result:
                user_id, employee_id, username, stored_hash, role = result

                # Check password using bcrypt
                if bcrypt.checkpw(password.encode(), stored_hash.encode()):
                    try:
                        # Try authenticating with Django's system
                        user = authenticate(request, username=employee_id, password=password)

                        # If not found in Django auth, create a user (sync with auth_user)
                        if user is None:
                            user = User.objects.filter(username=employee_id).first()
                            if not user:
                                user = User.objects.create_user(username=employee_id, password=password)
                            else:
                                user.set_password(password)
                                user.save()

                        # Log the user in (sets request.user and session)
                        login(request, user)

                        # Store any additional session data
                        request.session['role' ] = role
                        request.session['employee_id'] = employee_id
                        request.session['username'] = username

               

                        return JsonResponse({
                             'status': 'success',
                            'user_id': user_id,
                            'employee_id': employee_id,
                            'username': username,
                            'role': role
                        })

                    except Exception as e:
                        return JsonResponse({'status': 'error', 'message': f'Login failed: {str(e)}'})

                else:
                    return JsonResponse({'status': 'error', 'message': 'Incorrect password'})
            else:
                return JsonResponse({'status': 'error', 'message': 'User not found'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Server error: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})



def user_list(request):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, employee_id, username, email, role, phone_number
            FROM users
            ORDER BY id ASC
        """)
        rows = cursor.fetchall()

    users = []
    for row in rows:
        users.append({
            'id': row[0],
            'employee_id':row[1],
            'username': row[2],
            'email': row[3],
            'role': row[4],
            'phone': row[5]
        })

    return render(request, 'userinfo.html', {'users': users})

@csrf_exempt
def create_user(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            employee_id = data.get('employee_id', '').strip()
            username = data.get('username', '').strip()
            email = data.get('email', '').strip()
            password = data.get('password', '').strip()
            role = data.get('role', '').strip().lower()
            phone = data.get('phone', '').strip()

            # Field validation
            if not username or not email or not password or not role or not phone:
                return JsonResponse({'status': 'error', 'message': 'All fields are required'})

            # Phone number format check (Indian 10-digit, starts with 6-9)
            if not phone.isdigit() or len(phone) != 10 or not phone.startswith(('6', '7', '8', '9')):
                return JsonResponse({'status': 'error', 'message': 'Invalid phone number format'})

            # Check for duplicate username or phone
            with connection.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM users WHERE username = %s", [username])
                if cursor.fetchone()[0] > 0:
                    return JsonResponse({'status': 'error', 'message': 'Username already exists'})

                # cursor.execute("SELECT COUNT(*) FROM users WHERE phone_number = %s", [phone])
                # if cursor.fetchone()[0] > 0:
                #     return JsonResponse({'status': 'error', 'message': 'Phone number already in use'})

            # Hash password
            hashed_pw = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

            # Insert user
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO users ( employee_id, username, password, role, email, phone_number)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, [employee_id, username, hashed_pw, role, email, phone])

            return JsonResponse({'status': 'success', 'message': 'User created successfully'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@csrf_exempt
def edit_user(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_id = data.get('id')
            employee_id = data.get('employee_id').strip()
            username = data.get('username', '').strip()
            email = data.get('email', '').strip()
            role = data.get('role', '').strip().lower()
            phone = data.get('phone', '').strip()

            if not user_id or not username or not email or not role or not phone:
                return JsonResponse({'status': 'error', 'message': 'All fields are required'})

            # Optional: Duplicate check
            with connection.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM users WHERE username = %s AND id != %s", [username, user_id])
                if cursor.fetchone()[0] > 0:
                    return JsonResponse({'status': 'error', 'message': 'Username already exists'})

                cursor.execute("SELECT COUNT(*) FROM users WHERE phone_number = %s AND id != %s", [phone, user_id])
                if cursor.fetchone()[0] > 0:
                    return JsonResponse({'status': 'error', 'message': 'Phone number already in use'})

                cursor.execute("""
                    UPDATE users SET employee_id=%s, username=%s, email=%s, role=%s, phone_number=%s
                    WHERE id=%s
                """, [employee_id, username, email, role, phone, user_id])

            return JsonResponse({'status': 'success', 'message': 'User updated successfully'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})



@csrf_exempt
def logout_user(request):
    """Logout user and clear session (API endpoint)"""
    if request.method == 'POST':
        request.session.flush()
        return JsonResponse({'status': 'success', 'message': 'Logged out successfully'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@ensure_csrf_cookie
def check_session(request):
    """Check if user is logged in"""
    if request.session.get('is_authenticated'):
        return JsonResponse({
            'status': 'success',
            'username': request.session.get('username'),
            'user_type': request.session.get('user_type')
        })
    else:
        return JsonResponse({'status': 'error', 'message': 'Not authenticated'})

@ensure_csrf_cookie
def recover_username(request):
    """Recover username using employee ID (remove if employee_id column doesn't exist)"""
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id', '').strip()
        
        if not employee_id:
            return JsonResponse({
                'status': 'error', 
                'message': 'Employee ID is required'
            })
        
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT username FROM user WHERE employee_id = %s AND user_type = 'user'", 
                [employee_id]
            )
            result = cursor.fetchone()
        
        if result:
            return JsonResponse({
                'status': 'success', 
                'username': result[0]
            })
        else:
            return JsonResponse({
                'status': 'error', 
                'message': 'No user found with this Employee ID'
            })
    
    return JsonResponse({
        'status': 'error', 
        'message': 'Invalid request method'
    })

def dashboard(request):
    """Render the dashboard page with dynamic DC data"""
    username = request.session.get('username', '')
    if not username:
        return redirect('/')
    
    try:
        with connection.cursor() as cursor:
            # Open DCs
            cursor.execute("""
                SELECT COUNT(*) 
                FROM dc_details 
                WHERE created_by = %s AND status = 'PENDING'
            """, [username])
            open_dc_count = cursor.fetchone()[0]

            # Closed DCs (last 30 days)
            cursor.execute("""
                SELECT COUNT(*) 
                FROM dc_details 
                WHERE created_by = %s AND status = 'CLOSED'
                AND dc_date >= %s
            """, [username, (datetime.now() - timedelta(days=30)).date()])
            closed_dc_count = cursor.fetchone()[0]

            # Partial DCs
            cursor.execute("""
                SELECT COUNT(*) 
                FROM dc_details 
                WHERE created_by = %s AND status = 'PARTIAL'
            """, [username])
            partial_dc_count = cursor.fetchone()[0]

            # Total Revenue (from closed DCs in last 30 days)
            cursor.execute("""
                SELECT COALESCE(SUM(total_rate), 0) 
                FROM dc_details 
                WHERE created_by = %s 
                AND status = 'CLOSED' 
                AND dc_date >= %s
            """, [username, (datetime.now() - timedelta(days=30)).date()])
            total_revenue = cursor.fetchone()[0] or 0
            total_revenue = f"₹{float(total_revenue):,.2f}"

            # Weekly Report Data (Last 7 days DC counts)
            weekly_data = []
            for i in range(6, -1, -1):
                date = (datetime.now() - timedelta(days=i)).date()
                cursor.execute("""
                    SELECT COUNT(*) 
                    FROM dc_details 
                    WHERE created_by = %s 
                    AND DATE(dc_date) = %s
                """, [username, date])
                count = cursor.fetchone()[0]
                weekly_data.append({'date': date.strftime('%Y-%m-%d'), 'count': count})

            # Monthly Report Data (Last 30 days revenue)
            monthly_data = []
            for i in range(29, -1, -1):
                date = (datetime.now() - timedelta(days=i)).date()
                cursor.execute("""
                    SELECT COALESCE(SUM(total_rate), 0) 
                    FROM dc_details 
                    WHERE created_by = %s 
                    AND status = 'CLOSED' 
                    AND DATE(dc_date) = %s
                """, [username, date])
                revenue = cursor.fetchone()[0] or 0
                monthly_data.append({'date': date.strftime('%Y-%m-%d'), 'revenue': float(revenue)})

            # Recent Activities (DC-related)
            cursor.execute("""
                SELECT dc.dc_number, dc.party_name, dc.dc_date, dc.status, 
                       dcd.item_name, pdc.party_dc_number
                FROM dc_details dc
                LEFT JOIN dc_description dcd ON dc.dc_number = dcd.dc_number
                LEFT JOIN party_dc_details pdc ON dc.dc_number = pdc.dc_number
                WHERE dc.created_by = %s
                ORDER BY dc.updated_date_time DESC
                LIMIT 4
            """, [username])
            recent_activities = []
            for row in cursor.fetchall():
                dc_number, party_name, dc_date, status, item_name, party_dc_number = row
                activity_type = {
                    'PENDING': 'New DC Created',
                    'PARTIAL': 'Partial DC Updated',
                    'CLOSED': 'DC Delivered Successfully'
                }.get(status, 'Activity')
                recent_activities.append({
                    'dc_number': dc_number,
                    'party_name': party_name,
                    'item_name': item_name or 'N/A',
                    'party_dc_number': party_dc_number or 'N/A',
                    'dc_date': dc_date,
                    'status': status,
                    'type': activity_type,
                    'timestamp': dc_date.strftime('%Y-%m-%d %H:%M:%S')
                })

        return render(request, 'dashboard.html', {
            'username': username,
            'open_dc_count': open_dc_count,
            'closed_dc_count': closed_dc_count,
            'partial_dc_count': partial_dc_count,
            'total_revenue': total_revenue,
            'weekly_data': weekly_data,
            'monthly_data': monthly_data,
            'recent_activities': recent_activities
        })
    except Exception as e:
        logger.error(f"Error fetching dashboard data: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)

# [Rest of your views.py functions remain unchanged]

def dc_form(request):
    """Render the delivery challan form"""
    username = request.session.get('username', '')
    if not username:
        return redirect('/')
    return render(request, 'dc_form2.html', {
        'username': username,
        'view_mode': False,
        'states': [],
        'cities': []
    })
def report(request):
    """Render the report page"""
    return render(request, 'report.html')

def profile(request):
    """Render the profile page"""
    username = request.session.get('username', '')
    if not username:
        return redirect('/')
    return render(request, 'profile.html', {'username': username})


def logout(request):
    """Logout user and redirect to login page"""
    request.session.flush()
    return redirect('/')

def weasyprint_fetcher(url):
    """Custom URL fetcher for WeasyPrint"""
    try:
        logger.debug(f"Fetching URL: {url}")
        if url.startswith(django_settings.STATIC_URL):
            path = url.replace(django_settings.STATIC_URL, "").lstrip('/')
            static_file = os.path.join(django_settings.STATIC_ROOT, path).replace('\\', '/')
            if os.path.exists(static_file):
                mime_type = (
                    'image/png' if static_file.endswith('.png') else
                    'font/ttf' if static_file.endswith('.ttf') else
                    'application/octet-stream'
                )
                return {'file_obj': open(static_file, 'rb'), 'mime_type': mime_type}
            return default_url_fetcher(url)
        if hasattr(django_settings, 'MEDIA_URL') and url.startswith(django_settings.MEDIA_URL):
            path = url.replace(django_settings.MEDIA_URL, "").lstrip('/')
            media_file = os.path.join(django_settings.MEDIA_ROOT, path).replace('\\', '/')
            if os.path.exists(media_file):
                return {'file_obj': open(media_file, 'rb'), 'mime_type': 'application/octet-stream'}
            return default_url_fetcher(url)
        if url.startswith('file:///'):
            path = url.replace('file:///', '').lstrip('/')
            norm_path = os.path.normpath(path).replace('\\', '/')
            if os.path.exists(norm_path):
                mime_type = (
                    'image/png' if norm_path.endswith('.png') else
                    'font/ttf' if norm_path.endswith('.ttf') else
                    'application/octet-stream'
                )
                return {'file_obj': open(norm_path, 'rb'), 'mime_type': mime_type}
        return default_url_fetcher(url)
    except Exception as e:
        logger.error(f"Fetcher error for URL {url}: {e}", exc_info=True)
        return default_url_fetcher(url)

def generate_pdf_with_weasyprint(html_string):
    """Generate PDF using WeasyPrint"""
    try:
        buffer = BytesIO()
        static_root = os.path.normpath(str(django_settings.STATIC_ROOT)).replace('\\', '/')
        base_url = f"file:///{static_root}/"
        HTML(string=html_string, base_url=base_url, url_fetcher=weasyprint_fetcher).write_pdf(target=buffer)
        return buffer.getvalue()
    except Exception as e:
        logger.error(f"WeasyPrint error: {e}", exc_info=True)
        raise

@csrf_exempt
def get_next_dc_number(request):
    """Fetch the next globally unique DC number based on dc_type"""
    if request.method != 'GET':
        return JsonResponse({'message': 'Method not allowed'}, status=405)

    try:
        username = request.session.get('username')
        if not username:
            return JsonResponse({'message': 'User not authenticated'}, status=401)

        # Get dc_type from query parameters
        dc_type = request.GET.get('dc_type', '').upper()
        if dc_type not in ['SPM', 'VALVE_SECTION', 'QA']:
            return JsonResponse({'message': 'Invalid or missing dc_type'}, status=400)

        # Map dc_type to prefix and table name
        config_map = {
            'SPM': {
                'prefix': 'TLSPMDC',
                'table': 'spm_dc_numbers'
            },
            'VALVE_SECTION': {
                'prefix': 'TLVSDC', 
                'table': 'vs_dc_numbers'
            },
            'QA': {
                'prefix': 'TLQADC',
                'table': 'qa_dc_numbers'
            }
        }
        
        config = config_map[dc_type]
        prefix = config['prefix']
        table_name = config['table']

        # Get financial year
        current_year = datetime.now().year
        fy_start = str(current_year)[-2:]
        fy_end = str(current_year + 1)[-2:]
        fy = f"{fy_start}-{fy_end}"

        with transaction.atomic():
            with connection.cursor() as cursor:
                # Check if record exists for current financial year
                query = f"SELECT last_sequence FROM {table_name} WHERE financial_year = %s"
                cursor.execute(query, [fy])
                
                result = cursor.fetchone()
                
                if result:
                    # Update existing record
                    last_sequence = result[0]
                    next_sequence = last_sequence + 1
                    
                    update_query = f"UPDATE {table_name} SET last_sequence = %s WHERE financial_year = %s"
                    cursor.execute(update_query, [next_sequence, fy])
                else:
                    # Insert new record for this financial year
                    next_sequence = 1
                    insert_query = f"INSERT INTO {table_name} (financial_year, last_sequence) VALUES (%s, %s)"
                    cursor.execute(insert_query, [fy, next_sequence])

                # Generate the DC number
                next_dc = f"{prefix}-{fy}-{str(next_sequence).zfill(3)}"

        return JsonResponse({'dc_number': next_dc})

    except Exception as e:
        logger.error(f"Error fetching next DC number: {str(e)}")
        return JsonResponse({'message': f'Error: {str(e)}'}, status=500)

def delivery_challan_report(request):
    """Render delivery challan report for the authenticated user"""
    username = request.session.get('username', '')
    if not username:
        return redirect('/')
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT dc_number, party_name, dc_date
                FROM dc_details
                WHERE created_by = %s
                ORDER BY dc_date DESC
            """, [username])
            delivery_challans = [
                {
                    'dcNo': row[0],
                    'buyerName': row[1],
                    'dcDate': row[2]
                } for row in cursor.fetchall()
            ]
        return render(request, 'dc_report.html', {
            'username': username,
            'delivery_challans': delivery_challans
        })
    except Exception as e:
        logger.error(f"Error fetching delivery challan list for report: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)

def dictfetchall(cursor):
    """Return all rows from a cursor as a list of dictionaries"""
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def supplier_autocomplete(request):
    term = request.GET.get('term', '')
    
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, party_name, address_line_1, address_line_2, city, state, 
                       state_code, pincode, gstin_number
                FROM party_details 
                WHERE party_name LIKE %s 
                LIMIT 10
                """,
                [f'%{term}%']
            )
            columns = [col[0] for col in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Map database column names to expected field names
            formatted_results = []
            for row in results:
                formatted_data = {
                    'id': row['id'],
                    'party_name': row['party_name'],
                    'address1': row['address_line_1'],
                    'address2': row['address_line_2'],
                    'city': row['city'],
                    'state': row['state'],
                    'state_code': row['state_code'],
                    'pincode': row['pincode'],
                    'gstin': row['gstin_number']
                }
                formatted_results.append({
                    'label': row['party_name'],
                    'value': row['party_name'],
                    'data': formatted_data
                })
        
        return JsonResponse(formatted_results, safe=False)
        
    except Exception as e:
        logger.error(f"Supplier autocomplete error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

def get_states(request):
    term = request.GET.get('term', '')
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT DISTINCT state
                FROM party_details
                WHERE state IS NOT NULL AND state LIKE %s
                ORDER BY state
                """,
                [f'%{term}%']
            )
            states = [row[0] for row in cursor.fetchall()]
        return JsonResponse(states, safe=False)
    except Exception as e:
        logger.error(f"Error fetching states: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

def get_cities(request):
    state = request.GET.get('state', '')
    term = request.GET.get('term', '')
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT DISTINCT city
                FROM party_details
                WHERE city IS NOT NULL AND state = %s AND city LIKE %s
                ORDER BY city
                """,
                [state, f'%{term}%']
            )
            cities = [row[0] for row in cursor.fetchall()]
        return JsonResponse(cities, safe=False)
    except Exception as e:
        logger.error(f"Error fetching cities for state {state}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

def delivery_challan_view(request, dc_id):
    """View a specific delivery challan for the authenticated user"""

    # Ensure user is logged in
    username = request.session.get('username', '')
    if not username:
        logger.warning("Unauthorized access attempt to delivery challan view")
        return redirect('/')

    try:
        with connection.cursor() as cursor:
            # Fetch delivery challan + party details
            cursor.execute("""
                SELECT dc.dc_number, dc.party_name, dc.dc_date, dc.vehicle_no, dc.process,
                       pd.address_line_1, pd.address_line_2, pd.city, pd.state,
                       pd.state_code, pd.pincode, pd.gstin_number, dc.dc_type
                FROM dc_details dc
                JOIN party_details pd ON dc.party_name = pd.party_name
                WHERE dc.id = %s AND dc.created_by = %s
            """, [dc_id, username])
            dc_data = cursor.fetchone()

            if not dc_data:
                logger.warning(f"Delivery challan {dc_id} not found for user {username}")
                return HttpResponse("Delivery Challan not found or you don't have access", status=404)

            # Unpack the fetched values
            (dc_number, buyer_name, dc_date, vehicle_no, process, 
             address_line_1, address_line_2, city, state, state_code, 
             pincode, gstin, dc_type) = dc_data

            # Fetch item details with weight and square_feet
            cursor.execute("""
                SELECT id, item_name, description, uom, quantity, weight, square_feet, rate_per_each,
                       project_name, project_incharge, remarks, COALESCE(notes, '')
                FROM dc_description
                WHERE dc_number = %s
            """, [dc_number])

            items = [
                {
                    'id': row[0],
                    'item_name': row[1],
                    'description': row[2],
                    'uom': row[3],
                    'quantity': row[4],
                    'weight': row[5],
                    'square_feet': row[6],
                    'rate': row[7],
                    'project_name': row[8],
                    'project_incharge': row[9],
                    'remarks': row[10],
                    'notes': row[11],
                }
                for row in cursor.fetchall()
            ]

            # Fetch distinct states and cities
            cursor.execute("""
                SELECT DISTINCT state
                FROM party_details
                WHERE state IS NOT NULL
                ORDER BY state
            """)
            states = [row[0] for row in cursor.fetchall()]

            cursor.execute("""
                SELECT DISTINCT city
                FROM party_details
                WHERE city IS NOT NULL AND state = %s
                ORDER BY city
            """, [state])
            cities = [row[0] for row in cursor.fetchall()]

            # Assemble full context for template
            dc = {
                'dcNo': dc_number,
                'buyerName': buyer_name,
                'dcDate': dc_date,
                'vehicleNo': vehicle_no,
                'process': process,
                'dcType': dc_type,
                'buyerAddress1': address_line_1,
                'buyerAddress2': address_line_2,
                'buyerCity': city,
                'buyerState': state,
                'buyerStateCode': state_code,
                'buyerPincode': pincode,
                'buyerGstin': gstin,
                'items': items,
                'notes': items[0]['notes'] if items else '',
                'approxValue': ''
            }

        logger.info(f"Successfully fetched delivery challan {dc_id} for user {username}")
        return render(request, 'dc_form2.html', {
            'username': username,
            'dc': dc,
            'view_mode': True,
            'states': states,
            'cities': cities
        })

    except Exception as e:
        logger.error(f"Error fetching delivery challan {dc_id}: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)
    
def dc_item_details(request, item_id):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                id,
                ROW_NUMBER() OVER (ORDER BY id) as sno,
                item_name, description, uom, quantity,
                weight, weight_per_unit,
                square_feet, square_feet_per_unit, 
                rate_per_each, project_name,
                project_incharge, remarks
            FROM dc_description
            WHERE id = %s
        """, [item_id])
        item = dictfetchall(cursor)

    if not item:
        return JsonResponse({"error": "Item not found"}, status=404)

    return JsonResponse(item[0], safe=False)  # send single row



def supplier_autocomplete(request):
    term = request.GET.get('term', '')
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT party_name FROM party_details WHERE party_name LIKE %s", [f"%{term}%"])
        results = [{'label': row[0], 'value': row[0]} for row in cursor.fetchall()]
    return JsonResponse(results, safe=False)

def supplier_details(request):
    name = request.GET.get('name', '')
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT address_line_1, address_line_2, state, city, pincode, gstin_number, state_code
            FROM party_details WHERE party_name = %s
        """, [name])
        row = cursor.fetchone()

    if row:
        return JsonResponse({
            'address1': row[0],
            'address2': row[1],
            'state': row[2],
            'city': row[3],
            'pincode': row[4],
            'gstin': row[5],
            'state_code': row[6],
        })
    return JsonResponse({}, status=404)


def delivery_note_form(request):
    """Render the delivery note form"""
    return render(request, 'delivery_note_form.html')

def delivery_challan_json(request, dc_number):
    """Returns delivery challan data as JSON for a given dc_number"""
    username = request.session.get('username', '')
    if not username:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    try:
        with connection.cursor() as cursor:
            # Fetch header and buyer details
            cursor.execute("""
                SELECT dc.dc_number, dc.party_name, dc.dc_date, dc.vehicle_no, dc.process,
                       pd.address_line_1, pd.address_line_2, pd.city, pd.state,
                       pd.state_code, pd.pincode, pd.gstin_number
                FROM dc_details dc
                JOIN party_details pd ON dc.party_name = pd.party_name
                WHERE dc.dc_number = %s AND dc.created_by = %s
            """, [dc_number, username])
            dc_data = cursor.fetchone()

            if not dc_data:
                return JsonResponse({'error': 'Delivery Challan not found'}, status=404)

            (dc_number, buyer_name, dc_date, vehicle_no, process, address_line_1,
             address_line_2, city, state, state_code, pincode, gstin) = dc_data

            # Fetch items and notes (with COALESCE to avoid NULLs)
            cursor.execute("""
                SELECT item_name, description, uom, quantity, rate_per_each,
                       project_name, project_incharge, remarks, COALESCE(notes, '')
                FROM dc_description
                WHERE dc_number = %s
            """, [dc_number])
            items = [
                {
                    'item_name': row[0],
                    'description': row[1],
                    'uom': row[2],
                    'quantity': row[3],
                    'rate': row[4],
                    'project_name': row[5],
                    'mtr_dha': row[6],
                    'remarks': row[7],
                    'notes': row[8]
                } for row in cursor.fetchall()
            ]

            # Assemble final response object
            dc = {
                'dcNo': dc_number,
                'buyerName': buyer_name,
                'dcDate': str(dc_date),
                'vehicleNo': vehicle_no,
                'process': process,
                'buyerAddress1': address_line_1,
                'buyerAddress2': address_line_2,
                'buyerCity': city,
                'buyerState': state,
                'buyerStateCode': state_code,
                'buyerPincode': pincode,
                'buyerGstin': gstin,
                'items': items,
                'notes': items[0]['notes'] if items else '',  # Use first item's notes
                'approxValue': ''
            }

        return JsonResponse({'dc': dc}, status=200)

    except Exception as e:
        logger.error(f"Error fetching delivery challan JSON {dc_number}: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


def delivery_challan_list(request):
    username = request.session.get('username', '')
    if not username:
        return redirect('/')

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    dc.id,
                    dc.dc_number,
                    dc.dc_date,
                    dc.party_name,
                    dc.total_items,
                    dc.total_dispatch_quantity,
                    dc.process,
                    dc.vehicle_no,
                    dc.project_name
                FROM dc_details dc
                WHERE dc.created_by = %s AND dc.status = 'PENDING'
                ORDER BY dc.dc_date DESC
            """, [username])

            rows = cursor.fetchall()

        # build dc_list in one step
        dc_list = [
            {
                'id': row[0],
                'dcNo': row[1],
                'dcDate': row[2],
                'buyerName': row[3],
                'total_items': row[4],
                'total_dispatch_quantity': row[5],
                'process': row[6],
                'vehicleNo': row[7],
                'projectName': row[8]
            }
            for row in rows
        ]

        return render(request, 'delivery_challan_list.html', {
            'username': username,
            'delivery_challans': dc_list
        })

    except Exception as e:
        logger.error(f"Error fetching delivery challan list: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)

    
    
def partial_dc_list(request):
    """List delivery challans with status PARTIAL for the authenticated user"""
    username = request.session.get('username', '')
    if not username:
        return redirect('/')
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT dc.id, dc.dc_number, dc.party_name, dc.dc_date
                FROM dc_details dc
                WHERE dc.created_by = %s AND dc.status = 'PARTIAL'
                ORDER BY dc.dc_date DESC
            """, [username])
            delivery_challans = cursor.fetchall()
            dc_list = []
            for dc in delivery_challans:
                dc_id, dc_number, party_name, dc_date = dc
                cursor.execute("""
                    SELECT item_name
                    FROM dc_description
                    WHERE dc_number = %s
                """, [dc_number])
                items = [row[0] for row in cursor.fetchall() if row[0]]
                dc_list.append({
                    'id': dc_id,
                    'dcNo': dc_number,
                    'buyerName': party_name,
                    'dcDate': dc_date,
                    'items': items
                })
        return render(request, 'partial_dc.html', {
            'username': username,
            'delivery_challans': dc_list
        })
    except Exception as e:
        logger.error(f"Error fetching partial delivery challan list: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)
    
def close_dc_list(request):
    """List delivery challans with status CLOSED for the authenticated user"""
    username = request.session.get('username', '')
    if not username:
        return redirect('/')

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT dc.id, dc.dc_number, dc.party_name, dc.dc_date
                FROM dc_details dc
                WHERE dc.created_by = %s AND dc.status = 'CLOSED'
                ORDER BY dc.dc_date DESC
            """, [username])
            delivery_challans = cursor.fetchall()

            dc_list = []
            for dc in delivery_challans:
                dc_id, dc_number, party_name, dc_date = dc
                cursor.execute("""
                    SELECT item_name
                    FROM dc_description
                    WHERE dc_number = %s
                """, [dc_number])
                items = [row[0] for row in cursor.fetchall() if row[0]]
                dc_list.append({
                    'id': dc_id,
                    'dcNo': dc_number,
                    'buyerName': party_name,
                    'dcDate': dc_date,
                    'items': items
                })

        return render(request, 'close_dc.html', {
            'username': username,
            'delivery_chall': dc_list  # <-- FIX: match this in template
        })
    except Exception as e:
        logger.error(f"Error fetching closed delivery challan list: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)
    
    
@csrf_exempt
def update_delivery_note(request):
    """Update delivery note with received party and item details"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)

    try:
        username = request.session.get('username')
        if not username:
            return JsonResponse({'success': False, 'message': 'User not authenticated'}, status=401)

        data = json.loads(request.body)
        logger.debug(f"Received data: {json.dumps(data, indent=2)}")

        ist = pytz.timezone('Asia/Kolkata')
        current_datetime = datetime.now(ist)

        required_fields = ['dcNo', 'partyName', 'partyDcNumber', 'partyDcDate', 'received_items']
        for field in required_fields:
            if field not in data or data[field] in [None, '', []]:
                return JsonResponse({'success': False, 'message': f'Missing required field: {field}'}, status=400)

        dc_number = data['dcNo']
        party_name = data['partyName']
        party_dc_number = data['partyDcNumber']
        party_dc_date_str = data['partyDcDate']
        party_dc_doc_name = data.get('partyDcDocName', '')
        received_items = data['received_items']
        status = data.get('status', '').upper()

        if status not in ['CLOSED', 'PARTIAL']:
            return JsonResponse({'success': False, 'message': 'Invalid status provided'}, status=400)

        try:
            party_dc_date = datetime.strptime(party_dc_date_str, "%Y-%m-%d").date()
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid date format for partyDcDate. Use YYYY-MM-DD'}, status=400)

        with connection.cursor() as cursor:
            cursor.execute("SELECT id FROM party_details WHERE party_name = %s", [party_name])
            party = cursor.fetchone()
            if not party:
                return JsonResponse({'success': False, 'message': 'Party not found'}, status=404)

            cursor.execute("SELECT id FROM dc_details WHERE dc_number = %s AND created_by = %s", [dc_number, username])
            dc_row = cursor.fetchone()
            if not dc_row:
                return JsonResponse({'success': False, 'message': 'Delivery Challan not found'}, status=404)
            dc_id = dc_row[0]

            cursor.execute("SELECT COALESCE(SUM(quantity), 0) FROM dc_description WHERE dc_number = %s", [dc_number])
            total_dispatch_quantity = float(cursor.fetchone()[0])

        def safe_float(value, default=0.0):
            try:
                return float(value) if value is not None else default
            except (ValueError, TypeError):
                return default

        total_received_quantity = 0.0
        total_defect_quantity = 0.0

        for item in received_items:
            item_name = item.get('item_name')
            if not item_name:
                continue
            new_received_quantity = safe_float(item.get('received_quantity'))
            defect_quantity = safe_float(item.get('defect_quantity'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT quantity, received_quantity FROM dc_description
                    WHERE dc_number = %s AND item_name = %s
                """, [dc_number, item_name])
                result = cursor.fetchone()
                original_qty = safe_float(result[0]) if result else 0.0
                prev_received = safe_float(result[1]) if result else 0.0

            updated_received = prev_received + new_received_quantity if prev_received != 0.0 else new_received_quantity
            total_received_quantity += updated_received
            total_defect_quantity += defect_quantity

        if status == 'CLOSED' and abs(total_received_quantity + total_defect_quantity - total_dispatch_quantity) > 0.01:
            return JsonResponse({'success': False, 'message': 'Total received + defect must equal dispatch for CLOSED'}, status=400)
        elif status == 'PARTIAL' and (total_received_quantity + total_defect_quantity >= total_dispatch_quantity or total_received_quantity <= 0):
            return JsonResponse({'success': False, 'message': 'Total received + defect must be > 0 and < dispatch for PARTIAL'}, status=400)

        with transaction.atomic():
            with connection.cursor() as cursor:
                # Upsert into party_dc_details
                cursor.execute("SELECT id FROM party_dc_details WHERE dc_number = %s", [dc_number])
                if cursor.fetchone():
                    cursor.execute("""
                        UPDATE party_dc_details
                        SET party_dc_number = %s, party_dc_date = %s, party_dc_doc_name = %s, creation_date = %s
                        WHERE dc_number = %s
                    """, [party_dc_number, party_dc_date, party_dc_doc_name, current_datetime, dc_number])
                else:
                    cursor.execute("""
                        INSERT INTO party_dc_details (
                            dc_number, party_dc_number, party_dc_date, party_dc_doc_name, creation_date
                        ) VALUES (%s, %s, %s, %s, %s)
                    """, [dc_number, party_dc_number, party_dc_date, party_dc_doc_name, current_datetime])

                # Update item-wise quantities
                for item in received_items:
                    item_name = item.get('item_name')
                    if not item_name:
                        continue
                    remarks = item.get('remarks', '')
                    received_quantity = safe_float(item.get('received_quantity'))
                    defect_quantity = safe_float(item.get('defect_quantity'))
                    received_weight = safe_float(item.get('received_weight'), None)
                    new_original_quantity = safe_float(item.get('original_quantity'))

                    cursor.execute("""
                        SELECT quantity, received_quantity FROM dc_description
                        WHERE dc_number = %s AND item_name = %s
                    """, [dc_number, item_name])
                    result = cursor.fetchone()
                    prev_qty = safe_float(result[0]) if result else 0.0
                    prev_received = safe_float(result[1]) if result else 0.0

                    updated_qty = new_original_quantity
                    updated_received = prev_received + received_quantity if prev_received != 0.0 else received_quantity

                    cursor.execute("""
                        UPDATE dc_description
                        SET quantity = %s, received_quantity = %s, defect_quantity = %s, received_weight = %s, remarks = %s
                        WHERE dc_number = %s AND item_name = %s
                    """, [updated_qty, updated_received, defect_quantity, received_weight, remarks, dc_number, item_name])

                # Update dc_details
                cursor.execute("""
                    UPDATE dc_details
                    SET status = %s, updated_date_time = %s, total_received_quantity = %s
                    WHERE dc_number = %s AND created_by = %s
                """, [status, current_datetime, total_received_quantity, dc_number, username])

                # Insert into party_dc
                cursor.execute("""
                    INSERT INTO party_dc (
                        dc_number, dc_type, party_dc_number, party_dc_date, party_name, creation_date, received_quantity
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, [
                    dc_number,
                    status,
                    party_dc_number,
                    party_dc_date,
                    party_name,
                    current_datetime,
                    total_received_quantity
                ])

        return JsonResponse({
            'success': True,
            'message': f'Delivery note {dc_number} updated successfully with status {status}',
            'dc_id': dc_id
        })

    except Exception as e:
        logger.error(f"Error updating delivery note: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)



def partial_challan_view(request, dc_id):
    """Display read-only details of a partial delivery challan"""
    try:
        username = request.session.get('username')
        if not username:
            return HttpResponse("User not authenticated", status=401)

        # Fetch delivery challan details
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT dc.dc_number, dc.party_name, dc.dc_date, dc.status, dc.total_dispatch_quantity,
                       dc.total_received_quantity, pdc.party_dc_number, pdc.party_dc_date, pdc.party_dc_doc_name
                FROM dc_details dc
                LEFT JOIN party_dc_details pdc ON dc.dc_number = pdc.dc_number
                WHERE dc.id = %s AND dc.created_by = %s
            """, [dc_id, username])
            dc = cursor.fetchone()
            if not dc:
                return HttpResponse("Delivery Challan not found", status=404)

            dc_details = {
                'dc_number': dc[0],
                'party_name': dc[1],
                'dc_date': dc[2],
                'status': dc[3],
                'total_dispatch_quantity': float(dc[4]),
                'total_received_quantity': float(dc[5]) if dc[5] is not None else 0.0,
                'party_dc_number': dc[6] or '',
                'party_dc_date': dc[7] or '',
                'party_dc_doc_name': dc[8] or ''
            }

        # Fetch item details
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT item_name, quantity, received_quantity, remarks
                FROM dc_description
                WHERE dc_number = %s
            """, [dc_details['dc_number']])
            items = cursor.fetchall()
            item_list = [
                {
                    'item_name': item[0],
                    'dispatch_quantity': float(item[1]),
                    'partial_quantity': float(item[2]) if item[2] is not None else 0.0,
                    'remarks': item[3] or ''
                } for item in items
            ]

        # Calculate total partial quantity
        total_partial_quantity = sum(item['partial_quantity'] for item in item_list)

        context = {
            'dc': dc_details,
            'items': item_list,
            'total_partial_quantity': total_partial_quantity
        }
        return render(request, 'partial_view.html', context)

    except Exception as e:
        logger.error(f"Error fetching partial delivery challan details: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)
    
@csrf_exempt
def partial_update_note(request):
    """Update partial delivery note with received party and item details"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
        ist = pytz.timezone('Asia/Kolkata')
        current_datetime = datetime.now(ist)

        required_fields = ['dcNo', 'partyName', 'partyDcNumber', 'partyDcDate', 'received_items']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({'success': False, 'message': f'Missing required field: {field}'}, status=400)

        def safe_float(value, default=0.0):
            try:
                return float(value) if value is not None else default
            except (ValueError, TypeError):
                return default

        dc_number = data['dcNo']
        party_name = data['partyName']
        party_dc_number = data['partyDcNumber']
        party_dc_date = data['partyDcDate']
        party_dc_doc_name = data.get('partyDcDocName', '')
        received_items = data['received_items']
        status = data.get('status', '').upper()

        if status not in ['CLOSED', 'PARTIAL']:
            return JsonResponse({'success': False, 'message': 'Invalid status provided'}, status=400)

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id FROM party_details
                WHERE LOWER(party_name) = LOWER(%s)
            """, [party_name])
            party = cursor.fetchone()
            if not party:
                return JsonResponse({'success': False, 'message': 'Party not found'}, status=404)

            cursor.execute("""
                SELECT id, total_dispatch_quantity
                FROM dc_details
                WHERE dc_number = %s
            """, [dc_number])
            dc = cursor.fetchone()
            if not dc:
                return JsonResponse({'success': False, 'message': 'Delivery Challan not found'}, status=404)

            dc_id = dc[0]
            total_dispatch_quantity = safe_float(dc[1])

        total_received_qty = 0.0

        # Check and compute updated received quantities
        for item in received_items:
            item_name = item.get('item_name')
            new_received_quantity = safe_float(item.get('received_quantity'))
            dispatch_quantity = safe_float(item.get('dispatch_quantity'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT received_quantity FROM dc_description
                    WHERE dc_number = %s AND item_name = %s
                """, [dc_number, item_name])
                result = cursor.fetchone()
                prev_qty = safe_float(result[0]) if result else 0.0

            updated_total = prev_qty + new_received_quantity
            if updated_total > dispatch_quantity:
                return JsonResponse({'success': False, 'message': f'Received quantity for {item_name} exceeds dispatch quantity'}, status=400)

            total_received_qty += updated_total

        if status == 'CLOSED' and abs(total_received_qty - total_dispatch_quantity) > 0.01:
            return JsonResponse({'success': False, 'message': 'Total received must equal dispatch for CLOSED'}, status=400)
        elif status == 'PARTIAL' and total_received_qty >= total_dispatch_quantity:
            return JsonResponse({'success': False, 'message': 'Total received must be less than dispatch for PARTIAL'}, status=400)

        # Begin transaction for update
        with transaction.atomic():
            with connection.cursor() as cursor:
                # Upsert party_dc_details
                cursor.execute("SELECT id FROM party_dc_details WHERE dc_number = %s", [dc_number])
                if cursor.fetchone():
                    cursor.execute("""
                        UPDATE party_dc_details
                        SET party_dc_number = %s, party_dc_date = %s, party_dc_doc_name = %s, creation_date = %s
                        WHERE dc_number = %s
                    """, [party_dc_number, party_dc_date, party_dc_doc_name, current_datetime, dc_number])
                else:
                    cursor.execute("""
                        INSERT INTO party_dc_details (
                            dc_number, party_dc_number, party_dc_date, party_dc_doc_name, creation_date
                        ) VALUES (%s, %s, %s, %s, %s)
                    """, [dc_number, party_dc_number, party_dc_date, party_dc_doc_name, current_datetime])

                # Update each item's received quantity
                for item in received_items:
                    item_name = item.get('item_name')
                    new_received_quantity = safe_float(item.get('received_quantity'))
                    remarks = item.get('remarks', '')

                    cursor.execute("""
                        SELECT received_quantity FROM dc_description
                        WHERE dc_number = %s AND item_name = %s
                    """, [dc_number, item_name])
                    current_received = cursor.fetchone()
                    current_received_qty = safe_float(current_received[0]) if current_received else 0.0
                    updated_received_qty = current_received_qty + new_received_quantity

                    cursor.execute("""
                        UPDATE dc_description
                        SET received_quantity = %s, remarks = %s
                        WHERE dc_number = %s AND item_name = %s
                    """, [updated_received_qty, remarks, dc_number, item_name])

                # Update dc_details
                cursor.execute("""
                    UPDATE dc_details
                    SET status = %s, updated_date_time = %s, total_received_quantity = %s
                    WHERE dc_number = %s
                """, [status, current_datetime, total_received_qty, dc_number])

                # ✅ Insert into party_dc (new row for tracking every update)
                cursor.execute("""
                    INSERT INTO party_dc (
                        dc_number, dc_type, party_dc_number, party_dc_date, party_name, creation_date, received_quantity
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, [
                    dc_number,
                    status,
                    party_dc_number,
                    party_dc_date,
                    party_name,
                    current_datetime,
                    total_received_qty
                ])

        return JsonResponse({
            'success': True,
            'message': f'Delivery note {dc_number} updated successfully with status {status}',
            'dc_id': dc_id
        })

    except Exception as e:
        logger.error(f"Error updating partial delivery note: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)


def all_delivery_challan_list(request):
    """List all delivery challans (PENDING, PARTIAL, CLOSED) for the authenticated user"""
    username = request.session.get('username', '')
    if not username:
        return redirect('/')
    try:
        with connection.cursor() as cursor:
            # Fetch delivery challan details from dc_details
            cursor.execute("""
                SELECT dc.id, dc.dc_number, dc.party_name, dc.dc_date, dc.total_dispatch_quantity, dc.status
                FROM dc_details dc
                WHERE dc.created_by = %s
                ORDER BY dc.dc_date DESC
            """, [username])
            delivery_challans = cursor.fetchall()
            dc_list = []
            for dc in delivery_challans:
                dc_id, dc_number, party_name, dc_date, total_dispatch_quantity, status = dc
                # Fetch aggregated received_quantity from dc_description
                cursor.execute("""
                    SELECT COALESCE(SUM(received_quantity), 0) as received_quantity
                    FROM dc_description
                    WHERE dc_number = %s
                """, [dc_number])
                result = cursor.fetchone()
                received_quantity = float(result[0]) if result and result[0] is not None else 0.0
                # Fetch item names
                cursor.execute("""
                    SELECT item_name
                    FROM dc_description
                    WHERE dc_number = %s
                """, [dc_number])
                items = [row[0] for row in cursor.fetchall() if row[0]]
                dc_list.append({
                    'id': dc_id,
                    'dcNo': dc_number,
                    'buyerName': party_name,
                    'dcDate': dc_date,
                    'total_dispatch_quantity': float(total_dispatch_quantity),
                    'total_received_quantity': received_quantity,  # For template compatibility
                    'status': status,
                    'items': items
                })
        return render(request, 'all_dc.html', {
            'username': username,
            'delivery_challans': dc_list
        })
    except Exception as e:
        logger.error(f"Error fetching all delivery challan list: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)
    
    
def draft_delivery_challan(request):
    return render(request, 'draft.html', {
        'states': [],  # Replace with actual states list from indian_locations
        'cities': [],  # Replace with actual cities list
        'dc': {}       # Empty context for new drafts; populate if editing existing draft
    })

def draft_list(request):
    try:
        # Try session username first, fallback to request.user.username
        username = request.session.get('username')
        if not username and hasattr(request.user, 'username'):
            username = request.user.username
        if not username:
            logger.warning("No username found in session or user")
            return render(request, 'draft_list.html', {'drafts': [], 'error': 'User not authenticated'})

        logger.debug(f"Fetching drafts for user: {username}")
        with connection.cursor() as cursor:
            cursor.execute("""
           SELECT 
    d.id,
    d.party_name,
    d.process,
    COUNT(i.id) AS total_items,
    SUM(i.quantity) AS total_quantity,
    d.dc_type,
    d.vehicle_no,
    GROUP_CONCAT(DISTINCT i.project_name) AS project_names
FROM draft d
LEFT JOIN draft_items i ON d.id = i.draft_id
WHERE d.created_by = %s 
  AND d.status = 'DRAFT'
GROUP BY d.id, d.party_name, d.process, d.dc_type, d.vehicle_no
ORDER BY d.dc_date DESC;

""", [username])
            drafts = cursor.fetchall()
            logger.debug(f"Found {len(drafts)} drafts for user: {username}")

            draft_list = [
    {
        'id': row[0],
        'party_name': row[1],
        'process':row[2],
        'total_items': row[3],
        'total_quantity': row[4],
        'dc_type': row[5],
        'vehicle_no':row[6],
        'project_name': row[7]

    }
    for row in drafts
]

        return render(request, 'draft_list.html', {'drafts': draft_list})
    except Exception as e:
        logger.error(f"Error fetching draft list: {str(e)}", exc_info=True)
        return render(request, 'draft_list.html', {'drafts': [], 'error': str(e)})



def dictfetchall(cursor):
    """Return all rows from a cursor as a list of dictionaries."""
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def draft_view(request, draft_id):

    dc = None
    items = []

    try:
        with connection.cursor() as cursor:
            # Fetch draft (including all necessary fields)
            query = """
                SELECT 
                    d.id AS draft_id, d.dc_type AS dcType, d.dc_date AS dcDate,
                    d.party_name AS buyerName, d.vehicle_no AS vehicleNo, d.process,
                    d.total_dispatch_quantity, d.total_rate,
                    d.buyer_address1 AS buyerAddress1, d.buyer_address2 AS buyerAddress2,
                    d.buyer_city AS buyerCity, d.buyer_state AS buyerState,
                    d.buyer_state_code AS buyerStateCode, d.buyer_pincode AS buyerPincode,
                    d.buyer_gstin AS buyerGstin, d.notes AS notes, d.dc_number AS dcNo,
                    d.show_weight, d.show_square_feet
                FROM draft d
                WHERE d.id = %s
            """
            cursor.execute(query, [draft_id])
            dc_results = dictfetchall(cursor)

            if not dc_results:
                messages.error(request, f"No draft found with ID: {draft_id}")
                raise Http404("Draft not found")

            dc = dc_results[0]

            # Fetch draft items (including weight_per_unit and square_feet_per_unit)
            items_query = """
                SELECT 
                    id,
                    item_name, description, uom, quantity, weight, weight_per_unit,
                    square_feet, square_feet_per_unit, rate_per_each, project_name,
                    project_incharge, remarks
                FROM draft_items
                WHERE draft_id = %s
            """
            cursor.execute(items_query, [draft_id])
            items = dictfetchall(cursor)

    except Exception as e:
        messages.error(request, f"Error fetching draft: {str(e)}")
        raise Http404(f"Error fetching draft: {str(e)}")

    context = {
        'dc': dc,
        'items': items,
        'dc_notes': dc.get('notes', '')
    }

    return render(request, 'draft_view2.html', context)

def item_details(request, item_id):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                id,
                ROW_NUMBER() OVER (ORDER BY id) as sno,
                item_name, description, uom, quantity,
                weight, weight_per_unit,
                square_feet, square_feet_per_unit, 
                rate_per_each, project_name,
                project_incharge, remarks
            FROM draft_items
            WHERE id = %s
        """, [item_id])
        item = dictfetchall(cursor)

    if not item:
        return JsonResponse({"error": "Item not found"}, status=404)

    return JsonResponse(item[0], safe=False)  # send single row


@csrf_exempt
def get_next_dc_number(request):
    """Fetch the next globally unique DC number based on dc_type"""
    if request.method != 'GET':
        return JsonResponse({'message': 'Method not allowed'}, status=405)

    try:
        username = request.session.get('username')
        if not username:
            return JsonResponse({'message': 'User not authenticated'}, status=401)

        dc_type = request.GET.get('dc_type', '').upper()
        if dc_type not in ['SPM', 'VALVE_SECTION', 'QA']:
            return JsonResponse({'message': 'Invalid or missing dc_type'}, status=400)

        config_map = {
            'SPM': {'prefix': 'TLSPMDC', 'table': 'spm_dc_numbers'},
            'VALVE_SECTION': {'prefix': 'TLVSDC', 'table': 'valve_section_dc_numbers'},
            'QA': {'prefix': 'TLQADC', 'table': 'qa_dc_numbers'}
        }

        config = config_map[dc_type]
        prefix = config['prefix']
        table_name = config['table']

        current_year = datetime.now().year
        fy_start = str(current_year)[-2:]
        fy_end = str(current_year + 1)[-2:]
        fy = f"{fy_start}-{fy_end}"

        with transaction.atomic():
            with connection.cursor() as cursor:
                # LOCK the row during fetch
                query = f"SELECT last_sequence FROM {table_name} WHERE financial_year = %s FOR UPDATE"
                cursor.execute(query, [fy])
                result = cursor.fetchone()

                if result:
                    last_sequence = result[0]
                    next_sequence = last_sequence + 1

                    update_query = f"UPDATE {table_name} SET last_sequence = %s WHERE financial_year = %s"
                    cursor.execute(update_query, [next_sequence, fy])
                else:
                    next_sequence = 1
                    insert_query = f"INSERT INTO {table_name} (financial_year, last_sequence) VALUES (%s, %s)"
                    cursor.execute(insert_query, [fy, next_sequence])

                next_dc = f"{prefix}-{fy}-{str(next_sequence).zfill(3)}"

        return JsonResponse({'dc_number': next_dc})

    except Exception as e:
        logger.error(f"Error fetching next DC number: {str(e)}")
        return JsonResponse({'message': f'Error: {str(e)}'}, status=500)

def generate_pdf(request, data=None):
    """Generate PDF for delivery note"""
    username = request.session.get('username')
    if request.method == 'POST' and not data:
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'message': 'Invalid JSON'}, status=400)
    try:
        def format_date(date_str):
            if not date_str:
                return ''
            date = datetime.strptime(date_str, '%Y-%m-%d')
            return date.strftime('%d/%m/%Y')

        buyer_address = {
            'name': data.get('buyerName', ''),
            'address1': data.get('buyerAddress1', ''),
            'address2': data.get('buyerAddress2', ''),
            'city': data.get('buyerCity', ''),
            'pincode': data.get('buyerPincode', ''),
            'gstin': data.get('buyerGstin', ''),
            'state': data.get('buyerState', ''),
            'state_code': data.get('buyerStateCode', '')
        }

        items = data.get('items', [])

        # Calculate total quantity
        total_quantity = sum(float(item.get('quantity', 0)) for item in items if item.get('quantity'))

        for idx, item in enumerate(items, 1):
            item['sno'] = idx
            item['is_empty'] = False
            # Ensure name and description are present; if not, provide defaults
            item['name'] = item.get('item_name', '')
            item['description'] = item.get('description', '')
            item['quantity'] = item.get('quantity', '')
            
            
            


        items_per_page = 9
        last_page_items = 7
        pages = []

        if len(items) <= 7:
            # Special case: 3 or fewer items, first page gets all items and pads to 7
            page_items = items.copy()
            while len(page_items) < last_page_items:
                page_items.append({
                    'sno': '',
                    'name': '',
                    'description': '',
                    'uom': '',
                    'quantity': '',
                    'remarks': '',
                    'is_empty': True
                })
            pages.append(page_items)
        elif len(items) == 9:
            # Special case: exactly 9 items, first page gets 8, second page gets 1
            pages.append(items[:8].copy())
            pages.append(items[8:].copy())
            # Pad the last page to 7 items
            while len(pages[-1]) < last_page_items:
                pages[-1].append({
                    'sno': '',
                    'name': '',
                    'description': '',
                    'uom': '',
                    'quantity': '',
                    'remarks': '',
                    'is_empty': True
                })
        elif len(items) == 10:
            # Special case: exactly 10 items, first page gets 9, second page gets 1
            pages.append(items[:9].copy())
            pages.append(items[9:].copy())
            # Pad the last page to 7 items
            while len(pages[-1]) < last_page_items:
                pages[-1].append({
                    'sno': '',
                    'name': '',
                    'description': '',
                    'uom': '',
                    'quantity': '',
                    'remarks': '',
                    'is_empty': True
                })
        else:
            # General case: split items into pages
            for i in range(0, len(items), items_per_page):
                if i + items_per_page >= len(items) and len(items) > last_page_items:
                    # Last page should have up to 7 items
                    page_items = items[i:i + last_page_items].copy()
                    while len(page_items) < last_page_items:
                        page_items.append({
                            'sno': '',
                            'name': '',
                            'description': '',
                            'uom': '',
                            'quantity': '',
                            'remarks': '',
                            'is_empty': True
                        })
                else:
                    # Regular pages with 9 items
                    page_items = items[i:i + items_per_page].copy()
                    if len(page_items) < items_per_page and i + len(page_items) < len(items):
                        while len(page_items) < items_per_page:
                            page_items.append({
                                'sno': '',
                                'name': '',
                                'description': '',
                                'uom': '',
                                'quantity': '',
                                'remarks': '',
                                'is_empty': True
                            })
                pages.append(page_items)

        static_root = os.path.normpath(str(django_settings.STATIC_ROOT)).replace('\\', '/')
        base_file_url = f"file:///{static_root}/"
        
    
        with connection.cursor() as cursor:
            # Start a transaction to ensure atomicity
            with transaction.atomic():
                # 1. Fetch and copy dc_details to deleted_dc_details
                cursor.execute("""
                    SELECT weight_per_unit, weight
                    FROM dc_description
                    WHERE dc_number = %s
                """, [data.get('dcNo', '')])
                dc_description_rows = cursor.fetchone()
                print("dc_description_rows",dc_description_rows)
                weight_per_unit = dc_description_rows[0]
                weight = dc_description_rows[1]
                print(f"Weight per unit: {weight_per_unit}, Weight: {weight}")
                
                
        html_string = render_to_string('delivery_note_pdf.html', {
            'pages': list(enumerate(pages)),
            'buyer': buyer_address,
            'dc_no': data.get('dcNo', ''),
            'dc_date': format_date(data.get('dcDate', '')),
            'process': data.get('process', ''),
            'vehicle_no': data.get('vehicleNo', '') or '-',
            'notes': data.get('notes', ''),
            'approx_value': data.get('approxValue', ''),
            'total_quantity': total_quantity,
            'total_pages': len(pages),
            'teslead_logo': f"{base_file_url}images/tesleadd.png",
            'tuv_logo': f"{base_file_url}images/TUV INDIA 02.png",
            'inter_regular': f"{base_file_url}fonts/Inter_24pt-Regular.ttf",
            'inter_bold': f"{base_file_url} suburb/Inter_24pt-Bold.ttf",
            'username': username,
              'weight_per_unit': weight_per_unit,
    'weight': weight,
        })

        pdf_buffer = generate_pdf_with_weasyprint(html_string)
        dc_no = re.sub(r'[^\w\-]', '', str(data.get('dcNo', 'document')))
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Delivery_Note_{dc_no}.pdf"'
        response.write(pdf_buffer)
        return response

    except Exception as e:
        logger.error(f"Error generating PDF: {e}", exc_info=True)
        return JsonResponse({'message': f'Error generating PDF: {e}'}, status=400)


@csrf_exempt
def download_pdf(request):
    """Download PDF for delivery note"""
    if request.method != 'POST':
        return JsonResponse({'message': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body)
        return generate_pdf(request, data)
    except Exception as e:
        logger.error(f"Download PDF error: {e}", exc_info=True)
        return JsonResponse({'message': f'Error generating PDF: {e}'}, status=400)


@csrf_exempt
def save_delivery_note(request):
    """Save delivery note and remove draft entry"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
        dc_number = data.get('dcNumber')
        original_dc_number = data.get('original_dc_number', '')
        ist = pytz.timezone('Asia/Kolkata')
        current_datetime = datetime.now(ist)
        created_by = request.session.get('username', '')

        if not dc_number:
            return JsonResponse({'success': False, 'message': 'DC number is required'}, status=400)

        items = data.get('items', [])
        if not items:
            return JsonResponse({'success': False, 'message': 'At least one item is required'}, status=400)

        # Calculate total_quantity and total_rate
        total_quantity = sum(float(item.get('quantity', 0)) for item in items)
        total_rate = sum(float(item.get('quantity', 0)) * float(item.get('rate_per_each', 0)) for item in items)
        total_items = sum(1 for item in items if item.get('item_name'))  # ✅ count rows

        with transaction.atomic():
            with connection.cursor() as cursor:
                # If updating existing DC, delete old entries
                if original_dc_number:
                    cursor.execute("DELETE FROM dc_description WHERE dc_number = %s", [original_dc_number])
                    cursor.execute("DELETE FROM dc_details WHERE dc_number = %s", [original_dc_number])
                    logger.debug(f"Updated existing DC: {original_dc_number}")

                # Insert into final dc_details
                cursor.execute("""
                    INSERT INTO dc_details (
                        dc_date, dc_number, party_name, vehicle_no, process, dc_type,
                        total_dispatch_quantity, total_rate,total_items, project_name, project_incharge,
                        status, created_by, created_date_time, updated_date_time, show_weight, show_square_feet
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [
                    data.get('dcDate', ''),
                    dc_number,
                    data.get('buyerName', ''),
                    data.get('vehicleNo', ''),
                    data.get('process', ''),
                    data.get('dcType', 'SPM'),
                    total_quantity,
                    total_rate,
                    total_items,
                    items[0].get('project_name', '') if items else '',
                    items[0].get('project_incharge', '') if items else '',
                    'PENDING',
                    created_by,
                    current_datetime,
                    current_datetime,
                    data.get('showWeight', False),
                    data.get('showSquareFeet', False)
                ])

                # Insert into dc_description
                for item in items:
                    cursor.execute("""
                        INSERT INTO dc_description (
                            dc_number, item_name, description, uom, quantity, weight, square_feet,
                            rate_per_each, project_name, project_incharge, remarks, notes, created_by,
                            weight_per_unit, square_feet_per_unit
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, [
                        dc_number,
                        item.get('item_name', ''),
                        item.get('description', ''),
                        item.get('uom', ''),
                        float(item.get('quantity', 0)),
                        float(item.get('weight', 0)) if data.get('showWeight', False) else 0.0,
                        float(item.get('square_feet', 0)) if data.get('showSquareFeet', False) else 0.0,
                        float(item.get('rate_per_each', 0)),
                        item.get('project_name', ''),
                        item.get('project_incharge', ''),
                        item.get('remarks', ''),
                        data.get('notes', ''),
                        created_by,
                        float(item.get('weight_per_unit', 0)),
                        float(item.get('square_feet_per_unit', 0))
                    ])

                # Delete draft from draft and draft_items tables
                cursor.execute("DELETE FROM draft_items WHERE draft_id = %s", [data.get('draft_id')])
                cursor.execute("DELETE FROM draft WHERE id = %s", [data.get('draft_id')])
                logger.info(f"Draft deleted for DC number: {dc_number}, draft_id: {data.get('draft_id')}")

        return JsonResponse({
            'success': True,
            'message': 'Delivery note saved successfully and draft deleted',
            'dc_number': dc_number
        })

    except Exception as e:
        logger.error(f"Error saving delivery note: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)
    
    
def dictfetchall(cursor):
    """Convert raw SQL query results to a list of dictionaries."""
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def party_list_view(request):
    """Render the party list page."""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, party_name, address_line_1, address_line_2, city, state, state_code, 
                   pincode, gstin_number, email_id, phone_number
            FROM dc.party_details
            ORDER BY created_date DESC
        """)
        parties = dictfetchall(cursor)
    return render(request, 'party.html', {'parties': parties})

@csrf_exempt
def view_party(request, party_id):
    """Retrieve details of a specific party."""
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, party_name, address_line_1, address_line_2, city, state, state_code, 
                       pincode, gstin_number, email_id, phone_number
                FROM dc.party_details
                WHERE id = %s
            """, [party_id])
            party = cursor.fetchone()
            if not party:
                return JsonResponse({'status': 'error', 'message': 'Party not found'}, status=404)
            
            party_data = {
                'id': party[0],
                'party_name': party[1],
                'address_line_1': party[2],
                'address_line_2': party[3] or '',
                'city': party[4],
                'state': party[5],
                'state_code': party[6],
                'pincode': party[7],
                'gstin_number': party[8] or '',
                'email_id': party[9] or '',
                'phone_number': party[10] or ''
            }
        return JsonResponse({'status': 'success', 'party': party_data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@csrf_exempt
def create_party(request):
    """Create a new party in dc.party_details."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        party_name = data.get('party_name')
        address_line_1 = data.get('address_line_1')
        address_line_2 = data.get('address_line_2', '')
        city = data.get('city')
        state = data.get('state')
        state_code = data.get('state_code')
        pincode = data.get('pincode')
        gstin_number = data.get('gstin', '')
        email_id = data.get('email_id', '')
        phone_number = data.get('phone_number', '')
        created_by = request.user.username if not isinstance(request.user, AnonymousUser) else 'system'
        created_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Validation
        if not all([party_name, address_line_1, city, state, state_code, pincode]):
            return JsonResponse({'status': 'error', 'message': 'Required fields are missing'}, status=400)
        
        if not re.match(r'^\d{6}$', pincode):
            return JsonResponse({'status': 'error', 'message': 'Pincode must be a 6-digit number'}, status=400)
        
        if gstin_number and not re.match(r'^[0-9A-Z]{15}$', gstin_number):
            return JsonResponse({'status': 'error', 'message': 'Invalid GSTIN format'}, status=400)
        
        if email_id and not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email_id):
            return JsonResponse({'status': 'error', 'message': 'Invalid email format'}, status=400)
        
        if phone_number and not re.match(r'^\d{10}$', phone_number):
            return JsonResponse({'status': 'error', 'message': 'Phone number must be a 10-digit number'}, status=400)

        with connection.cursor() as cursor:
            # Check if party_name already exists
            cursor.execute("SELECT id FROM dc.party_details WHERE party_name = %s", [party_name])
            if cursor.fetchone():
                return JsonResponse({'status': 'error', 'message': 'Party name already exists'}, status=400)
            
            # Insert new party
            cursor.execute("""
                INSERT INTO dc.party_details (
                    party_name, address_line_1, address_line_2, city, state, state_code, pincode, 
                    gstin_number, email_id, phone_number, created_by, created_date, updated_by, updated_date
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, [
                party_name, address_line_1, address_line_2 or None, city, state, state_code, pincode,
                gstin_number or None, email_id or None, phone_number or None, created_by, created_date,
                created_by, created_date
            ])
        
        return JsonResponse({'status': 'success', 'message': 'Party created successfully'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@csrf_exempt
def edit_party(request):
    """Update an existing party in dc.party_details."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        party_id = data.get('id')
        party_name = data.get('party_name')
        address_line_1 = data.get('address_line_1')
        address_line_2 = data.get('address_line_2', '')
        city = data.get('city')
        state = data.get('state')
        state_code = data.get('state_code')
        pincode = data.get('pincode')
        gstin_number = data.get('gstin', '')
        email_id = data.get('email_id', '')
        phone_number = data.get('phone_number', '')
        updated_by = request.user.username if not isinstance(request.user, AnonymousUser) else 'system'
        updated_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Validation
        if not all([party_id, party_name, address_line_1, city, state, state_code, pincode]):
            return JsonResponse({'status': 'error', 'message': 'Required fields are missing'}, status=400)
        
        if not re.match(r'^\d{6}$', pincode):
            return JsonResponse({'status': 'error', 'message': 'Pincode must be a 6-digit number'}, status=400)
        
        if gstin_number and not re.match(r'^[0-9A-Z]{15}$', gstin_number):
            return JsonResponse({'status': 'error', 'message': 'Invalid GSTIN format'}, status=400)
        
        if email_id and not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email_id):
            return JsonResponse({'status': 'error', 'message': 'Invalid email format'}, status=400)
        
        if phone_number and not re.match(r'^\d{10}$', phone_number):
            return JsonResponse({'status': 'error', 'message': 'Phone number must be a 10-digit number'}, status=400)

        with connection.cursor() as cursor:
            # Check if party exists
            cursor.execute("SELECT id FROM dc.party_details WHERE id = %s", [party_id])
            if not cursor.fetchone():
                return JsonResponse({'status': 'error', 'message': 'Party not found'}, status=404)
            
            # Check if party_name is taken by another party
            cursor.execute("SELECT id FROM dc.party_details WHERE party_name = %s AND id != %s", [party_name, party_id])
            if cursor.fetchone():
                return JsonResponse({'status': 'error', 'message': 'Party name already exists'}, status=400)
            
            # Update party
            cursor.execute("""
                UPDATE dc.party_details
                SET party_name = %s, address_line_1 = %s, address_line_2 = %s, city = %s, state = %s,
                    state_code = %s, pincode = %s, gstin_number = %s, email_id = %s, phone_number = %s,
                    updated_by = %s, updated_date = %s
                WHERE id = %s
            """, [
                party_name, address_line_1, address_line_2 or None, city, state, state_code, pincode,
                gstin_number or None, email_id or None, phone_number or None, updated_by, updated_date, party_id
            ])
        
        return JsonResponse({'status': 'success', 'message': 'Party updated successfully'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@csrf_exempt
def delete_party(request):
    """Delete a party from dc.party_details."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        party_id = data.get('id')

        with connection.cursor() as cursor:
            # Check if party exists
            cursor.execute("SELECT id FROM dc.party_details WHERE id = %s", [party_id])
            if not cursor.fetchone():
                return JsonResponse({'status': 'error', 'message': 'Party not found'}, status=404)
            
            # Delete party
            cursor.execute("DELETE FROM dc.party_details WHERE id = %s", [party_id])
        
        return JsonResponse({'status': 'success', 'message': 'Party deleted successfully'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    




def item_list_view(request):
    """Render the item list page."""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, item_name, description, uom, quantity, rate_per_each, project_name, 
                   project_incharge, remarks
            FROM dc.dc_description
            ORDER BY created_date DESC
        """)
        items = dictfetchall(cursor)
        
    return render(request, 'items.html', {'items': items})

@csrf_exempt
def view_item(request, item_id):
    """Retrieve details of a specific item."""
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, item_name, description, uom, quantity, rate_per_each, project_name, 
                       project_incharge, remarks,weight,square_feet
                FROM dc.dc_description
                WHERE id = %s
            """, [item_id])
            item = cursor.fetchone()
            if not item:
                return JsonResponse({'status': 'error', 'message': 'Item not found'}, status=404)
            
        item_data = {
        'id': item[0],
        'item_name': item[1],
        'description': item[2] or '',
        'uom': item[3],
        'quantity': item[4],
        'rate_per_each': item[5],
        'project_name': item[6] or '',
        'project_incharge': item[7] or '',
        'remarks': item[8] or '',
        'weight': item[9],           # Add weight
        'square_feet': item[10]      # Add square feet
    }

        return JsonResponse({'status': 'success', 'item': item_data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    


@csrf_exempt
def create_item(request):
    """Create a new item in dc.dc_description."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        print("Received data:", data)  # Debug log

        item_name = data.get('item_name')
        description = data.get('description', '')
        uom = data.get('uom')
        quantity = data.get('quantity')
        rate_per_each = data.get('rate_per_each')
        project_name = data.get('project_name', '')
        project_incharge = data.get('project_incharge', '')
        remarks = data.get('remarks', '')

        # Convert square_feet and weight safely
        try:
            square_feet = float(data.get('square_feet', 0))
        except (ValueError, TypeError):
            return JsonResponse({'status': 'error', 'message': 'Square Feet must be a valid number'}, status=400)

        try:
            weight = float(data.get('weight', 0))
        except (ValueError, TypeError):
            return JsonResponse({'status': 'error', 'message': 'Weight must be a valid number'}, status=400)

        created_by = request.user.username if not isinstance(request.user, AnonymousUser) else 'system'
        created_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Required fields check
        if not all([item_name, uom, quantity, rate_per_each]):
            return JsonResponse({'status': 'error', 'message': 'Required fields are missing'}, status=400)

        # Validate quantity
        try:
            quantity = int(quantity)
            if quantity < 0:
                return JsonResponse({'status': 'error', 'message': 'Quantity must be a non-negative number'}, status=400)
        except (TypeError, ValueError):
            return JsonResponse({'status': 'error', 'message': 'Quantity must be a valid number'}, status=400)

        # Validate rate
        try:
            rate_per_each = float(rate_per_each)
            if rate_per_each < 0:
                return JsonResponse({'status': 'error', 'message': 'Rate Per Each must be a non-negative number'}, status=400)
        except (TypeError, ValueError):
            return JsonResponse({'status': 'error', 'message': 'Rate Per Each must be a valid number'}, status=400)

        # Validate UOM
        valid_uoms = ['KG', 'NOS', 'LTR', 'MTR', 'SQFT', 'CUM', 'TON', 'SET', 'BAG', 'BOX']
        if uom not in valid_uoms:
            return JsonResponse({'status': 'error', 'message': f'Invalid UOM. Must be one of {", ".join(valid_uoms)}'}, status=400)

        with connection.cursor() as cursor:
            # Check duplicate item_name
            cursor.execute("SELECT id FROM dc.dc_description WHERE item_name = %s", [item_name])
            if cursor.fetchone():
                return JsonResponse({'status': 'error', 'message': 'Item name already exists'}, status=400)

            # Debug log before insert
            insert_values = [
                item_name, description or None, uom, quantity, square_feet, weight,
                rate_per_each, project_name or None, project_incharge or None, remarks or None,
                created_by, created_date, created_by, created_date
            ]
            print("Insert values:", insert_values)  # Debug log

            # Perform insert
            cursor.execute("""
                INSERT INTO dc.dc_description (
                    item_name, description, uom, quantity, square_feet, weight,
                    rate_per_each, project_name, project_incharge, remarks,
                    created_by, created_date, updated_by, updated_date
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, insert_values)

        return JsonResponse({'status': 'success', 'message': 'Item created successfully'})

    except Exception as e:
        print("Error:", str(e))  # Log to console for debugging
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@csrf_exempt
def edit_item(request):
    """Update an existing item in dc.dc_description."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        item_id = data.get('id')
        item_name = data.get('item_name')
        description = data.get('description', '')
        uom = data.get('uom')
        quantity = data.get('quantity')
        square_feet = data.get('square_feet', 0)
        weight = data.get('weight', 0)
        rate_per_each = data.get('rate_per_each')
        project_name = data.get('project_name', '')
        project_incharge = data.get('project_incharge', '')
        remarks = data.get('remarks', '')
        updated_by = request.user.username if not isinstance(request.user, AnonymousUser) else 'system'
        updated_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Validation
        if not all([item_id, item_name, uom, quantity, rate_per_each]):
            return JsonResponse({'status': 'error', 'message': 'Required fields are missing'}, status=400)
        
        try:
            quantity = int(quantity)
            if quantity < 0:
                return JsonResponse({'status': 'error', 'message': 'Quantity must be a non-negative number'}, status=400)
        except (TypeError, ValueError):
            return JsonResponse({'status': 'error', 'message': 'Quantity must be a valid number'}, status=400)
        
        try:
            rate_per_each = float(rate_per_each)
            if rate_per_each < 0:
                return JsonResponse({'status': 'error', 'message': 'Rate Per Each must be a non-negative number'}, status=400)
        except (TypeError, ValueError):
            return JsonResponse({'status': 'error', 'message': 'Rate Per Each must be a valid number'}, status=400)
        
        valid_uoms = ['KG', 'NOS', 'LTR', 'MTR', 'SQFT', 'CUM', 'TON', 'SET', 'BAG', 'BOX']
        if uom not in valid_uoms:
            return JsonResponse({'status': 'error', 'message': f'Invalid UOM. Must be one of {", ".join(valid_uoms)}'}, status=400)

        with connection.cursor() as cursor:
            # Check if item exists
            cursor.execute("SELECT id FROM dc.dc_description WHERE id = %s", [item_id])
            if not cursor.fetchone():
                return JsonResponse({'status': 'error', 'message': 'Item not found'}, status=404)
            
            # Check if item_name is taken by another item
            cursor.execute("SELECT id FROM dc.dc_description WHERE item_name = %s AND id != %s", [item_name, item_id])
            if cursor.fetchone():
                return JsonResponse({'status': 'error', 'message': 'Item name already exists'}, status=400)
            
            # Update item
            cursor.execute("""
                UPDATE dc.dc_description
                SET item_name = %s, description = %s, uom = %s, quantity = %s, rate_per_each = %s,
                    project_name = %s, project_incharge = %s, remarks = %s,
                    updated_by = %s, updated_date = %s, weight = %s, square_feet = %s
                WHERE id = %s
            """, [
                item_name, description or None, uom, quantity, square_feet, weight, rate_per_each, project_name or None,
                project_incharge or None, remarks or None, updated_by, updated_date, item_id
            ])
        
        return JsonResponse({'status': 'success', 'message': 'Item updated successfully'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@csrf_exempt
def delete_item(request):
    """Delete an item from dc.dc_description."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        item_id = data.get('id')

        with connection.cursor() as cursor:
            # Check if item exists
            cursor.execute("SELECT id FROM dc.dc_description WHERE id = %s", [item_id])
            if not cursor.fetchone():
                return JsonResponse({'status': 'error', 'message': 'Item not found'}, status=404)
            
            # Delete item
            cursor.execute("DELETE FROM dc.dc_description WHERE id = %s", [item_id])
        
        return JsonResponse({'status': 'success', 'message': 'Item deleted successfully'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
def settings_page(request):
    return render(request, 'settings.html')

def get_authenticated_user(request):
    """Helper function to get the authenticated user ID from the session."""
    if not request.user.is_authenticated:
        return None, JsonResponse({'status': 'error', 'message': 'User not authenticated'}, status=401)
    return request.user.id, None

@csrf_exempt
def get_user_info(request):
    if request.method == 'GET':
        try:
            user_id, error_response = get_authenticated_user(request)
            if error_response:
                return error_response

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT username, email, role, phone_number
                    FROM users
                    WHERE id = %s
                """, [user_id])
                user = cursor.fetchone()

                if not user:
                    return JsonResponse({'status': 'error', 'message': 'User not found'}, status=404)

                return JsonResponse({
                    'status': 'success',
                    'data': {
                        'username': user[0],
                        'email': user[1],
                        'role': user[2],
                        'phone_number': user[3] if user[3] else ''
                    }
                })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def update_profile(request):
    if request.method == 'POST':
        try:
            user_id, error_response = get_authenticated_user(request)
            if error_response:
                return error_response

            data = json.loads(request.body)
            username = data.get('username', '').strip()
            email = data.get('email', '').strip()
            phone_number = data.get('phone_number', '').strip()

            # Validate inputs
            if not username or not email:
                return JsonResponse({'status': 'error', 'message': 'Username and email are required'}, status=400)

            # Validate phone number if provided
            if phone_number and (not phone_number.isdigit() or len(phone_number) != 10 or not phone_number.startswith(('6', '7', '8', '9'))):
                return JsonResponse({'status': 'error', 'message': 'Invalid phone number format'}, status=400)

            # Check for duplicate username or email (excluding current user)
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT COUNT(*) FROM users 
                    WHERE (username = %s OR email = %s) AND id != %s
                """, [username, email, user_id])
                if cursor.fetchone()[0] > 0:
                    return JsonResponse({'status': 'error', 'message': 'Username or email already in use'}, status=400)

                # Check for duplicate phone number if provided
                if phone_number:
                    cursor.execute("SELECT COUNT(*) FROM users WHERE phone_number = %s AND id != %s", [phone_number, user_id])
                    if cursor.fetchone()[0] > 0:
                        return JsonResponse({'status': 'error', 'message': 'Phone number already in use'}, status=400)

                # Update user profile
                cursor.execute("""
                    UPDATE users 
                    SET username = %s, email = %s, phone_number = %s
                    WHERE id = %s
                """, [username, email, phone_number or None, user_id])

            return JsonResponse({'status': 'success', 'message': 'Profile updated successfully'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def change_password(request):
    if request.method == 'POST':
        try:
            user_id, error_response = get_authenticated_user(request)
            if error_response:
                return error_response

            data = json.loads(request.body)
            current_password = data.get('current_password', '').strip()
            new_password = data.get('new_password', '').strip()

            if not current_password or not new_password:
                return JsonResponse({'status': 'error', 'message': 'All fields are required'}, status=400)

            # Validate new password strength (example: min 8 chars, 1 letter, 1 number)
            if len(new_password) < 8 or not any(c.isalpha() for c in new_password) or not any(c.isdigit() for c in new_password):
                return JsonResponse({'status': 'error', 'message': 'New password must be at least 8 characters long and contain letters and numbers'}, status=400)

            with connection.cursor() as cursor:
                # Verify current password
                cursor.execute("SELECT password FROM users WHERE id = %s", [user_id])
                stored_password = cursor.fetchone()

                if not stored_password or not bcrypt.checkpw(current_password.encode(), stored_password[0].encode()):
                    return JsonResponse({'status': 'error', 'message': 'Current password is incorrect'}, status=400)

                # Hash new password
                hashed_new_password = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()

                # Update password
                cursor.execute("UPDATE users SET password = %s WHERE id = %s", [hashed_new_password, user_id])

            return JsonResponse({'status': 'success', 'message': 'Password updated successfully'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def send_otp(request):
    if request.method == 'POST':
        try:
            # Replace with your authentication logic
            user_id, error_response = get_authenticated_user(request)
            if error_response:
                return error_response

            data = json.loads(request.body)
            email = data.get('email', '').strip()

            with connection.cursor() as cursor:
                cursor.execute("SELECT username, email FROM users WHERE id = %s", [user_id])
                result = cursor.fetchone()

            if not result or result[1] != email:
                return JsonResponse({'status': 'error', 'message': 'Invalid email address'}, status=400)

            username = result[0]
            department = "IT Department"

            # Generate new OTP
            otp = ''.join(random.choices(string.digits, k=6))
            created_at = datetime.now()
            expires_at = created_at + timedelta(minutes=1)  # 1-minute expiry

            with connection.cursor() as cursor:
                # Delete any existing OTPs for this user
                cursor.execute("DELETE FROM user_otp WHERE user_id = %s", [user_id])
                
                # Insert new OTP
                cursor.execute("""
                    INSERT INTO user_otp (user_id, otp, created_at, expires_at)
                    VALUES (%s, %s, %s, %s)
                """, [user_id, otp, created_at, expires_at])

            # HTML email with logo
            with open("static/images/Teslead-Logo-White.png", "rb") as image_file:
             encoded = base64.b64encode(image_file.read()).decode()
            logo_url = f"data:image/png;base64,{encoded}"


            html_content = f"""
                <html>
                    <body style="font-family: Arial, sans-serif; padding: 20px; background-color: #f4f4f4;">
                        <div style="max-width: 500px; margin: auto; background: white; padding: 20px; border-radius: 8px;">
                            <div style="text-align: center;">
                                <img src="{logo_url}" alt="Teslead Logo" style="max-width: 150px; margin-bottom: 20px;">
                            </div>
                            <h2 style="color: #333;">Hello, {username.title()}</h2>
                            <p>You have requested to reset your password from the <strong>{department}</strong>.</p>
                            <h1 style="text-align: center; color: #007bff;">{otp}</h1>
                            <p>This OTP is valid for 1 minute.</p>
                            <hr>
                            <p style="font-size: 12px; color: #888;">If you did not request this, please ignore this email.</p>
                        </div>
                    </body>
                </html>
            """

            # Send email
            try:
                msg = MIMEMultipart('alternative')
                msg['From'] = settings.EMAIL_HOST_USER
                msg['To'] = email
                msg['Subject'] = 'Your OTP for Password Reset'
                msg.attach(MIMEText(html_content, 'html'))

                server = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
                server.starttls()
                server.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)
                server.sendmail(settings.EMAIL_HOST_USER, email, msg.as_string())
                server.quit()

                return JsonResponse({'status': 'success', 'message': 'OTP sent to your email'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': 'Failed to send OTP: ' + str(e)}, status=500)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def change_password_otp(request):
    if request.method == 'POST':
        try:
            user_id, error_response = get_authenticated_user(request)
            if error_response:
                return error_response

            data = json.loads(request.body)
            otp = data.get('otp', '').strip()
            new_password = data.get('new_password', '').strip()

            with connection.cursor() as cursor:
                # Delete expired OTPs
                cursor.execute("DELETE FROM user_otp WHERE expires_at < %s", [datetime.now()])

                # Verify OTP
                cursor.execute("""
                    SELECT id, expires_at 
                    FROM user_otp 
                    WHERE user_id = %s AND otp = %s
                    ORDER BY created_at DESC LIMIT 1
                """, [user_id, otp])
                otp_record = cursor.fetchone()

                if not otp_record:
                    return JsonResponse({'status': 'error', 'message': 'Invalid OTP'}, status=400)

                otp_id, expires_at = otp_record
                if datetime.now() > expires_at:
                    return JsonResponse({'status': 'error', 'message': 'OTP has expired'}, status=400)

                # If no new password provided, this is just OTP verification
                if not new_password:
                    return JsonResponse({'status': 'success', 'message': 'OTP verified successfully'})

                # Validate new password
                if len(new_password) < 8 or not any(c.isalpha() for c in new_password) or not any(c.isdigit() for c in new_password):
                    return JsonResponse({'status': 'error', 'message': 'New password must be at least 8 characters long and contain letters and numbers'}, status=400)

                # Hash new password
                hashed_new_password = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()

                # Update password
                cursor.execute("UPDATE users SET password = %s WHERE id = %s", [hashed_new_password, user_id])

                # Delete used OTP
                cursor.execute("DELETE FROM user_otp WHERE id = %s", [otp_id])

            return JsonResponse({'status': 'success', 'message': 'Password updated successfully'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


def supplier_autocomplete(request):
    term = request.GET.get('term', '')
    
    # Add error handling and debugging
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, party_name, address_line_1, address_line_2, city, state, 
                       state_code, pincode, gstin_number
                FROM party_details 
                WHERE party_name LIKE %s 
                LIMIT 10
                """,
                [f'%{term}%']
            )
            columns = [col[0] for col in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Map database column names to expected field names
            formatted_results = []
            for row in results:
                formatted_data = {
                    'id': row['id'],
                    'party_name': row['party_name'],
                    'address1': row['address_line_1'],  # Map address_line_1 to address1
                    'address2': row['address_line_2'],  # Map address_line_2 to address2
                    'city': row['city'],
                    'state': row['state'],
                    'state_code': row['state_code'],
                    'pincode': row['pincode'],
                    'gstin': row['gstin_number']  # Map gstin_number to gstin
                }
                formatted_results.append({
                    'label': row['party_name'],
                    'value': row['party_name'],
                    'data': formatted_data
                })
        
        return JsonResponse(formatted_results, safe=False)
        
    except Exception as e:
        # Log the error for debugging
        print(f"Supplier autocomplete error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


def item_autocomplete(request):
    term = request.GET.get('term', '')
    
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, item_name, description, uom, quantity,weight,square_feet rate_per_each, 
                       project_name,project_incharge, remarks,square_feet,weight
                FROM dc_description 
                WHERE item_name LIKE %s 
                LIMIT 10
                """,
                [f'%{term}%']
            )
            columns = [col[0] for col in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Map database column names to expected field names
            formatted_results = []
            for row in results:
                formatted_data = {
                    'id': row['id'],
                    'item_name': row['item_name'],
                    'description': row['description'],
                    'uom': row['uom'],
                    'quantity': row['quantity'],
                    'square_feet': row['square_feet'],  # This column doesn't exist in your table
                    'weight': row['weight'],  # This column doesn't exist in your table
                    'rate': row['rate_per_each'],  # Map rate_per_each to rate
                    'project_name': row['project_name'],
                    'project_incharge': row['project_incharge'], # This column doesn't exist in your table
                    'remarks': row['remarks']
                    
                }
                formatted_results.append({
                    'label': row['item_name'],
                    'value': row['item_name'],
                    'data': formatted_data
                })
        
        return JsonResponse(formatted_results, safe=False)
        
    except Exception as e:
        # Log the error for debugging
        print(f"Item autocomplete error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)
    
    
def dc_description_autocomplete(request):
    term = request.GET.get('term', '')
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT dc_no, dc_type, dc_date, process, vehicle_no, notes 
            FROM dc_description 
            WHERE dc_no LIKE %s OR process LIKE %s LIMIT 10
            """,
            [f'%{term}%', f'%{term}%']
        )
        columns = [col[0] for col in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        formatted_results = [{'label': row['dc_no'], 'value': row['dc_no'], 'data': row} for row in results]
    return JsonResponse(formatted_results, safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def add_supplier(request):
    try:
        # Log the raw request data
        logger.info(f"Request method: {request.method}")
        logger.info(f"Request body: {request.body}")
        logger.info(f"Content type: {request.content_type}")
        
        # Parse JSON data
        try:
            data = json.loads(request.body)
            logger.info(f"Parsed data: {data}")
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {e}")
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON data.'}, status=400)
        
        # Get current user and timestamp
        current_user = request.user.username if request.user.is_authenticated else 'Anonymous'
        current_time = timezone.now()
        
        logger.info(f"User: {current_user}, Time: {current_time}")
        
        # Extract and validate data
        party_name = data.get('name', '').strip()
        address1 = data.get('address1', '').strip()
        address2 = data.get('address2', '').strip()
        state = data.get('state', '').strip()
        city = data.get('city', '').strip()
        pincode = data.get('pincode', '').strip()
        gstin = data.get('gstin', '').strip()
        state_code = data.get('state_code', '').strip()
        
        logger.info(f"Extracted data - Name: {party_name}, State: {state}, City: {city}")
        
        # Validate required fields
        if not all([party_name, address1, state, city, pincode, state_code]):
            missing_fields = []
            if not party_name: missing_fields.append('name')
            if not address1: missing_fields.append('address1')
            if not state: missing_fields.append('state')
            if not city: missing_fields.append('city')
            if not pincode: missing_fields.append('pincode')
            if not state_code: missing_fields.append('state_code')
            
            logger.warning(f"Missing required fields: {missing_fields}")
            return JsonResponse({
                'status': 'error', 
                'message': f'Missing required fields: {", ".join(missing_fields)}'
            }, status=400)
        
        # Validate pincode
        if not (pincode.isdigit() and len(pincode) == 6):
            logger.warning(f"Invalid pincode: {pincode}")
            return JsonResponse({'status': 'error', 'message': 'Invalid pincode format.'}, status=400)
        
        # Validate GSTIN if provided
        if gstin and not (
            len(gstin) == 15 and
            gstin[:2].isdigit() and
            gstin[2:7].isalpha() and
            gstin[7:11].isdigit() and
            gstin[11:12].isalpha() and
            gstin[12:13] in '1234567890ABCDEFGHIJKLMNOPQRSTUVWXYZ' and
            gstin[13:14] == 'Z' and
            gstin[14:15].isalnum()
        ):
            logger.warning(f"Invalid GSTIN: {gstin}")
            return JsonResponse({'status': 'error', 'message': 'Invalid GSTIN format.'}, status=400)
        
        # Database operations
        try:
            with connection.cursor() as cursor:
                # Check if supplier exists
                logger.info("Checking if supplier exists...")
                cursor.execute("SELECT COUNT(*) FROM party_details WHERE party_name = %s", [party_name])
                count = cursor.fetchone()[0]
                logger.info(f"Existing supplier count: {count}")
                
                if count > 0:
                    return JsonResponse({'status': 'error', 'message': 'Supplier name already exists.'}, status=400)
                
                # Insert new supplier with audit fields
                logger.info("Inserting new supplier...")
                cursor.execute(
                    """
                    INSERT INTO party_details (party_name, address_line_1, address_line_2, state, city, pincode, gstin_number, state_code, created_by, created_date, updated_date)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    [party_name, address1, address2 or None, state, city, pincode, gstin or None, state_code, current_user, current_time, current_time]
                )
                
                # Get the inserted supplier ID
                supplier_id = cursor.lastrowid
                logger.info(f"Supplier inserted successfully with ID: {supplier_id}")
                
        except Exception as db_error:
            logger.error(f"Database error: {db_error}")
            return JsonResponse({
                'status': 'error', 
                'message': f'Database error: {str(db_error)}'
            }, status=500)
        
        return JsonResponse({
            'status': 'success', 
            'message': f'Supplier added successfully by {current_user}.',
            'supplier_id': supplier_id,
            'created_by': current_user,
            'created_date': current_time.strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        logger.error(f"Unexpected error in add_supplier: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error', 
            'message': f'Server error: {str(e)}'
        }, status=500)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import connection
import json

@csrf_exempt
def check_supplier(request):
    if request.method == "POST":
        data = json.loads(request.body)
        name = data.get("name", "").strip()

        print("🔎 Incoming supplier name:", repr(name))  # debug

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, party_name, address_line_1, address_line_2, state, city, pincode, gstin_number, state_code
                FROM party_details
                WHERE TRIM(LOWER(party_name)) = TRIM(LOWER(%s))
            """, [name])
            row = cursor.fetchone()

        if row:
            supplier = {
                "id": row[0],
                "name": row[1],
                "address1": row[2],
                "address2": row[3],
                "state": row[4],
                "city": row[5],
                "pincode": row[6],
                "gstin": row[7],
                "state_code": row[8],
            }
            print("✅ Supplier found:", supplier)   # debug
            return JsonResponse({"exists": True, "supplier": supplier})

        else:
            print("❌ No supplier found for:", name)   # debug
            return JsonResponse({"exists": False})

    return JsonResponse({"error": "Invalid request"}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def update_supplier(request):
    try:
        data = json.loads(request.body)
        supplier_id = data.get('supplier_id')
        
        if not supplier_id:
            return JsonResponse({'status': 'error', 'message': 'Supplier ID is required.'}, status=400)
        
        # Get current user and timestamp
        current_user = request.user.username if request.user.is_authenticated else 'Anonymous'
        current_time = timezone.now()
        
        # Extract and validate data
        party_name = data.get('name', '').strip()
        address1 = data.get('address1', '').strip()
        address2 = data.get('address2', '').strip()
        state = data.get('state', '').strip()
        city = data.get('city', '').strip()
        pincode = data.get('pincode', '').strip()
        gstin = data.get('gstin', '').strip()
        state_code = data.get('state_code', '').strip()
        
        # Validate required fields
        if not all([party_name, address1, state, city, pincode, state_code]):
            return JsonResponse({'status': 'error', 'message': 'All required fields must be filled.'}, status=400)
        
        # Validate pincode
        if not (pincode.isdigit() and len(pincode) == 6):
            return JsonResponse({'status': 'error', 'message': 'Invalid pincode format.'}, status=400)
        
        # Validate GSTIN if provided
        if gstin and not (
            len(gstin) == 15 and
            gstin[:2].isdigit() and
            gstin[2:7].isalpha() and
            gstin[7:11].isdigit() and
            gstin[11:12].isalpha() and
            gstin[12:13] in '1234567890ABCDEFGHIJKLMNOPQRSTUVWXYZ' and
            gstin[13:14] == 'Z' and
            gstin[14:15].isalnum()
        ):
            return JsonResponse({'status': 'error', 'message': 'Invalid GSTIN format.'}, status=400)
        
        try:
            with connection.cursor() as cursor:
                # Check if supplier exists and get current data
                cursor.execute("SELECT party_name, created_by, created_date FROM party_details WHERE id = %s", [supplier_id])
                result = cursor.fetchone()
                
                if not result:
                    return JsonResponse({'status': 'error', 'message': 'Supplier not found.'}, status=404)
                
                original_name, created_by, created_date = result
                
                # Check if another supplier with the same name exists (excluding current one)
                cursor.execute("SELECT COUNT(*) FROM party_details WHERE party_name = %s AND id != %s", [party_name, supplier_id])
                if cursor.fetchone()[0] > 0:
                    return JsonResponse({'status': 'error', 'message': 'Supplier name already exists.'}, status=400)
                
                # Update supplier (keeping original created_by and created_date)
                cursor.execute(
                    """
                    UPDATE party_details 
                    SET party_name = %s, address_line_1 = %s, address_line_2 = %s, state = %s, 
                        city = %s, pincode = %s, gstin_number = %s, state_code = %s, updated_date = %s
                    WHERE id = %s
                    """,
                    [party_name, address1, address2 or None, state, city, pincode, gstin or None, state_code, current_time, supplier_id]
                )
                
        except Exception as db_error:
            return JsonResponse({'status': 'error', 'message': f'Database error: {str(db_error)}'}, status=500)
        
        return JsonResponse({
            'status': 'success', 
            'message': f'Supplier updated successfully by {current_user}.',
            'updated_by': current_user,
            'updated_date': current_time.strftime('%Y-%m-%d %H:%M:%S'),
            'created_by': created_by,
            'created_date': created_date.strftime('%Y-%m-%d %H:%M:%S') if created_date else None
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Server error: {str(e)}'}, status=500)


@require_http_methods(["GET"])
def get_supplier_details(request, supplier_id):
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT party_name, address_line_1, address_line_2, state, city, pincode, 
                       gstin_number, state_code, created_by, created_date, updated_date
                FROM party_details 
                WHERE id = %s
                """, 
                [supplier_id]
            )
            result = cursor.fetchone()
            
            if not result:
                return JsonResponse({'status': 'error', 'message': 'Supplier not found.'}, status=404)
            
            supplier_data = {
                'name': result[0],
                'address1': result[1],
                'address2': result[2],
                'state': result[3],
                'city': result[4],
                'pincode': result[5],
                'gstin': result[6],
                'state_code': result[7],
                'created_by': result[8],
                'created_date': result[9].strftime('%Y-%m-%d %H:%M:%S') if result[9] else None,
                'updated_date': result[10].strftime('%Y-%m-%d %H:%M:%S') if result[10] else None
            }
            
            return JsonResponse({'status': 'success', 'data': supplier_data})
            
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Server error: {str(e)}'}, status=500)
    
    
@csrf_exempt
def save_draft(request):
    if request.method == 'POST':
        try:
            # Parse JSON data from request body
            data = json.loads(request.body)

            # Extract data from request
            dc_date = data.get('dcDate')
            dc_type = data.get('dcType')
            dc_number = data.get('dcNo', '')
            process = data.get('process')
            vehicle_no = data.get('vehicleNo')
            party_name = data.get('buyerName')
            # created_by = request.user.username if request.user.is_authenticated else 'system'
            created_by = request.session.get('username') or 'system'
            created_by_id = request.session.get('employee_id', 'system')
            created_time = updated_time = datetime.now(pytz.UTC)
            buyer_address1 = data.get('buyerAddress1')
            buyer_address2 = data.get('buyerAddress2')
            buyer_state = data.get('buyerState')
            buyer_city = data.get('buyerCity')
            buyer_pincode = data.get('buyerPincode')
            buyer_gstin = data.get('buyerGstin')
            buyer_state_code = data.get('buyerStateCode')
            notes = data.get('notes')
            show_weight = data.get('showWeight', True)
            show_square_feet = data.get('showSquareFeet', True)
            items = data.get('items', [])

            # Insert draft into database
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO draft (
                        dc_date, dc_number, party_name, vehicle_no, process, dc_type,
                        created_by, created_by_id, created_date_time, updated_date_time, notes,
                        buyer_address1, buyer_address2, buyer_state, buyer_city, buyer_pincode,
                        buyer_gstin, buyer_state_code, show_weight, show_square_feet
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s
                    )
                """, [
                    dc_date, dc_number, party_name, vehicle_no, process, dc_type,
                    created_by, created_by_id, created_time, updated_time, notes,
                    buyer_address1, buyer_address2, buyer_state, buyer_city, buyer_pincode,
                    buyer_gstin, buyer_state_code, show_weight, show_square_feet
                ])

                # ✅ MySQL-compatible way to get the last inserted ID
                draft_id = cursor.lastrowid

                # Insert items into draft_items table
                for item in items:
                    cursor.execute("""
                        INSERT INTO draft_items (
                            draft_id, item_name, description, uom, quantity, rate_per_each, remarks,  
                            project_name, weight, square_feet, project_incharge, weight_per_unit, 
                            square_feet_per_unit
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, [
                            draft_id,
                            item.get('item_name'),
                            item.get('description'),
                            item.get('uom'),
                            item.get('quantity'),
                            item.get('rate_per_each'),
                            item.get('remarks'),
                            item.get('project_name'),
                            item.get('weight', 0),
                            item.get('square_feet', 0),
                            item.get('project_incharge'),
                            item.get('weight_per_unit', 0),
                            item.get('square_feet_per_unit', 0)
                    ])

            return JsonResponse({'success': True, 'draft_id': draft_id})
            
 
        

        except Exception as e:
            # Log the full traceback to console
            print(traceback.format_exc())
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)



@csrf_exempt
def update_draft(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            draft_id = data.get('draft_id')
            if not draft_id:
                return JsonResponse({'success': False, 'message': 'Draft ID is required'}, status=400)

            # Extract draft fields
            dc_date = data.get('dcDate')
            dc_type = data.get('dcType')
            dc_number = data.get('dcNo', '')
            process = data.get('process')
            vehicle_no = data.get('vehicleNo')
            party_name = data.get('buyerName')
            updated_time = datetime.now(pytz.UTC)
            buyer_address1 = data.get('buyerAddress1')
            buyer_address2 = data.get('buyerAddress2')
            buyer_state = data.get('buyerState')
            buyer_city = data.get('buyerCity')
            buyer_pincode = data.get('buyerPincode')
            buyer_gstin = data.get('buyerGstin')
            buyer_state_code = data.get('buyerStateCode')
            notes = data.get('notes')
            items = data.get('items', [])

            # Validate required fields
            required_fields = ['dcDate', 'dcType', 'process', 'buyerName', 'buyerAddress1', 
                             'buyerState', 'buyerCity', 'buyerPincode', 'buyerStateCode']
            for field in required_fields:
                if not data.get(field):
                    return JsonResponse({
                        'success': False, 
                        'message': f'Missing required field: {field}'
                    }, status=400)

            # Validate items
            if not items:
                return JsonResponse({
                    'success': False, 
                    'message': 'At least one item is required'
                }, status=400)

            # Calculate show_weight and show_square_feet based on item data
            show_weight = any(
                float(item.get('weight_per_unit', 0)) > 0 or float(item.get('weight', 0)) > 0
                for item in items
            )
            show_square_feet = any(
                float(item.get('square_feet_per_unit', 0)) > 0 or float(item.get('square_feet', 0)) > 0
                for item in items
            )

            with connection.cursor() as cursor:
                # Update draft table
                cursor.execute("""
                    UPDATE draft
                    SET dc_date = %s, dc_number = %s, party_name = %s, vehicle_no = %s,
                        process = %s, dc_type = %s, updated_date_time = %s, notes = %s,
                        buyer_address1 = %s, buyer_address2 = %s, buyer_state = %s,
                        buyer_city = %s, buyer_pincode = %s, buyer_gstin = %s,
                        buyer_state_code = %s, show_weight = %s, show_square_feet = %s
                    WHERE id = %s
                """, [
                    dc_date, dc_number, party_name, vehicle_no, process, dc_type,
                    updated_time, notes, buyer_address1, buyer_address2, buyer_state,
                    buyer_city, buyer_pincode, buyer_gstin, buyer_state_code,
                    show_weight, show_square_feet, draft_id
                ])
                
                # Check if draft exists
                # if not cursor.fetchone():
                if cursor.rowcount == 0:
                    return JsonResponse({
                        'success': False, 
                        'message': 'Draft not found'
                    }, status=404)

                # Delete existing items
                cursor.execute("DELETE FROM draft_items WHERE draft_id = %s", [draft_id])

                # Insert updated items
                for item in items:
                    # Ensure numeric fields are valid
                    quantity = float(item.get('quantity', 0)) if item.get('quantity') is not None else 0
                    weight = float(item.get('weight', 0)) if item.get('weight') is not None else 0
                    weight_per_unit = float(item.get('weight_per_unit', 0)) if item.get('weight_per_unit') is not None else 0
                    square_feet = float(item.get('square_feet', 0)) if item.get('square_feet') is not None else 0
                    square_feet_per_unit = float(item.get('square_feet_per_unit', 0)) if item.get('square_feet_per_unit') is not None else 0
                    rate_per_each = float(item.get('rate_per_each', 0)) if item.get('rate_per_each') is not None else 0

                    cursor.execute("""
                        INSERT INTO draft_items (
                            draft_id, item_name, description, uom, quantity, weight, 
                            weight_per_unit, square_feet, square_feet_per_unit, rate_per_each,
                            remarks, project_name, project_incharge
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, [
                        draft_id,
                        item.get('item_name', ''),
                        item.get('description', ''),
                        item.get('uom', ''),
                        quantity,
                        weight,
                        weight_per_unit,
                        square_feet,
                        square_feet_per_unit,
                        rate_per_each,
                        item.get('remarks', ''),
                        item.get('project_name', ''),
                        item.get('project_incharge', '')
                    ])

            return JsonResponse({
                'success': True, 
                'draft_id': draft_id,
                'showWeight': show_weight,
                'showSquareFeet': show_square_feet
            })

        except Exception as e:
            print(f"Error in update_draft: {str(e)}")
            print(traceback.format_exc())
            return JsonResponse({
                'success': False, 
                'message': f'Error updating draft: {str(e)}'
            }, status=500)



    return JsonResponse({
        'success': False, 
        'message': 'Invalid request method'
    }, status=400)



@csrf_exempt
def delete_draft(request, draft_id):
    if request.method == "POST":
        try:
            with connection.cursor() as cursor:

                cursor.execute("DELETE FROM draft_items WHERE draft_id = %s", [draft_id])

                cursor.execute("DELETE FROM draft WHERE id = %s", [draft_id])  
                # replace `draft_table` with your actual table name

            return JsonResponse({"success": True, "message": "Draft deleted successfully!"})
        except Exception as e:
            print(traceback.format_exc())
            return JsonResponse({"success": False, "message": str(e)}, status=500)

    return JsonResponse({"success": False, "message": "Invalid request method"}, status=400)



  


from django.http import JsonResponse
from django.db import connection
from django.views.decorators.csrf import csrf_exempt
import traceback

@csrf_exempt
def get_draft(request, draft_id):
    try:
        with connection.cursor() as cursor:
            # Fetch draft details
            cursor.execute("""
                SELECT dc_date, dc_number, party_name, vehicle_no, process, dc_type,
                       notes, buyer_address1, buyer_address2, buyer_state, buyer_city,
                       buyer_pincode, buyer_gstin, buyer_state_code, show_weight, show_square_feet
                FROM draft
                WHERE id = %s
            """, [draft_id])
            draft = cursor.fetchone()
            if not draft:
                return JsonResponse({'success': False, 'message': 'Draft not found'}, status=404)

            draft_data = {
                'dcDate': draft[0].strftime('%Y-%m-%d') if draft[0] else '',
                'dcNo': draft[1] or '',
                'buyerName': draft[2] or '',
                'vehicleNo': draft[3] or '',
                'process': draft[4] or '',
                'dcType': draft[5] or '',
                'notes': draft[6] or '',
                'buyerAddress1': draft[7] or '',
                'buyerAddress2': draft[8] or '',
                'buyerState': draft[9] or '',
                'buyerCity': draft[10] or '',
                'buyerPincode': draft[11] or '',
                'buyerGstin': draft[12] or '',
                'buyerStateCode': draft[13] or '',
                'showWeight': bool(draft[14]),
                'showSquareFeet': bool(draft[15])
            }

            # Fetch draft items
            cursor.execute("""
                SELECT item_name, description, uom, quantity, weight, weight_per_unit, 
                       square_feet, square_feet_per_unit, rate_per_each, remarks, 
                       project_name, project_incharge
                FROM draft_items
                WHERE draft_id = %s
            """, [draft_id])
            items = cursor.fetchall()
            items_data = [
                {
                    'item_name': item[0] or '',
                    'description': item[1] or '',
                    'uom': item[2] or '',
                    'quantity': float(item[3]) if item[3] is not None else 0,
                    'weight': float(item[4]) if item[4] is not None else 0,
                    'weight_per_unit': float(item[5]) if item[5] is not None else 0,
                    'square_feet': float(item[6]) if item[6] is not None else 0,
                    'square_feet_per_unit': float(item[7]) if item[7] is not None else 0,
                    'rate_per_each': float(item[8]) if item[8] is not None else 0,
                    'remarks': item[9] or '',
                    'project_name': item[10] or '',
                    'project_incharge': item[11] or ''
                } for item in items
            ]

            draft_data['items'] = items_data
            return JsonResponse({'success': True, 'data': draft_data})

    except Exception as e:
        print(f"Error in get_draft: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'message': f'Error fetching draft: {str(e)}'}, status=500)






from django.shortcuts import render, Http404
from django.contrib import messages
from django.db import connection

def dictfetchall(cursor):
    """Return all rows from a cursor as a dict"""
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def closed_dc_view(request, dc_id):
    dc = None
    items = []
    received_items = []
    party_dc_numbers = []
    dc_notes = ""
    party_dc_details = []
    
    with connection.cursor() as cursor:
        # Fetch DC details (from dc_details table)
        query = """
            SELECT 
                d.id AS dc_id, 
                d.dc_type AS dcType, 
                d.dc_date AS dcDate,
                d.party_name AS buyerName, 
                d.party_name AS partyName,
                d.vehicle_no AS vehicleNo, 
                d.process,
                d.total_dispatch_quantity, 
                d.total_rate, 
                d.project_name,
                d.project_incharge, 
                d.dc_number AS dcNo,
                p.address_line_1 AS buyerAddress1, 
                p.address_line_2 AS buyerAddress2,
                p.city AS buyerCity, 
                p.state AS buyerState, 
                p.state_code AS buyerStateCode,
                p.pincode AS buyerPincode, 
                p.gstin_number AS buyerGstin
            FROM dc_details d
            LEFT JOIN party_details p ON d.party_name = p.party_name
            WHERE d.id = %s AND d.status = 'CLOSED'
        """
        cursor.execute(query, [dc_id])
        dc_results = dictfetchall(cursor)
        
        if not dc_results:
            messages.error(request, f"No closed DC found with ID: {dc_id}")
            raise Http404("Closed DC not found")
        
        dc = dc_results[0]
        
        if dc['dcNo']:
            # Fetch DC items (from dc_description table)
            items_query = """
                SELECT 
                    dd.item_name, 
                    dd.description, 
                    dd.uom, 
                    dd.quantity, 
                    dd.weight, 
                    dd.square_feet,
                    dd.rate_per_each, 
                    dd.project_name, 
                    dd.project_incharge, 
                    dd.remarks, 
                    dd.notes,
                    pdd.party_dc_number
                FROM dc_description dd
                LEFT JOIN party_dc_details pdd ON dd.dc_number = pdd.dc_number
                WHERE dd.dc_number = %s
                ORDER BY dd.id
            """
            cursor.execute(items_query, [dc['dcNo']])
            items = dictfetchall(cursor)
            
            # Fetch notes from dc_description table
            notes_query = """
                SELECT notes
                FROM dc_description
                WHERE dc_number = %s AND notes IS NOT NULL AND notes != ''
                ORDER BY id
            """
            cursor.execute(notes_query, [dc['dcNo']])
            notes_results = cursor.fetchall()
            if notes_results:
                dc_notes = notes_results[0][0]
            
            # Fetch Party DC Numbers (from party_dc_details table)
            party_dc_query = """
                SELECT DISTINCT party_dc_number
                FROM party_dc_details
                WHERE dc_number = %s
                ORDER BY party_dc_number
            """
            cursor.execute(party_dc_query, [dc['dcNo']])
            party_dc_numbers = [row[0] for row in cursor.fetchall()]
            
            # Fetch Party DC Details for the table
            party_dc_details_query = """
                SELECT 
                    dc_number,
                    party_dc_number,
                    party_dc_date,
                    dc_type AS status,
                    creation_date
                FROM party_dc
                WHERE dc_number = %s
                ORDER BY party_dc_number
            """
            cursor.execute(party_dc_details_query, [dc['dcNo']])
            party_dc_details = dictfetchall(cursor)
            
            # Debug: Print party DC details
            print(f"Party DC Details: {party_dc_details}")
            
            # Fetch Received Items
            received_items_query = """
                SELECT 
                    dd.item_name, 
                    dd.description, 
                    dd.uom, 
                    dd.received_quantity,
                    dd.received_weight AS weight, 
                    dd.square_feet,
                    dd.rate_per_each, 
                    dd.project_name, 
                    dd.project_incharge, 
                    dd.remarks,
                    dd.updated_date AS received_date,
                    pdd.party_dc_number
                FROM dc_description dd
                LEFT JOIN party_dc_details pdd ON dd.dc_number = pdd.dc_number
                WHERE dd.dc_number = %s AND dd.received_quantity IS NOT NULL
                ORDER BY dd.id
            """
            cursor.execute(received_items_query, [dc['dcNo']])
            received_items = dictfetchall(cursor)
    
    # Prepare party DC numbers for the template
    dc['partyDcNumbers'] = party_dc_numbers
    if dc['dcType'] == 'Direct' and party_dc_numbers:
        dc['partyDcNo'] = party_dc_numbers[0]
    
    context = {
        'dc': dc,
        'items': items,
        'received_items': received_items,
        'dc_notes': dc_notes,
        'party_dc_details': party_dc_details
    }
    
    return render(request, 'closed_dc_view.html', context)



def overall_dc(request):
    try:
        # Get username from session or user object
        username = request.session.get('username')
        if not username and hasattr(request.user, 'username'):
            username = request.user.username
        if not username:
            logger.warning("No username found in session or user")
            return render(request, 'overall_dc.html', {'overalldc': [], 'error': 'User not authenticated'})

        logger.debug(f"Fetching overall DC for user: {username}")

        with connection.cursor() as cursor:
            # Fetch all records joined on dc_number
            cursor.execute("""
                SELECT 
                    dd.dc_number,
                    dd.dc_type,
                    dd.party_name,
                    dd.status,
                    ddesc.item_name,
                    ddesc.created_date,
                    ddesc.created_by
                    
                FROM dc_description ddesc
                JOIN dc_details dd ON ddesc.dc_number = dd.dc_number
                ORDER BY dd.dc_number DESC
            """)
            rows = cursor.fetchall()

        # Group data by dc_number
        dc_map = {}
        for row in rows:
            dc_number = row[0]
            if dc_number not in dc_map:
                dc_map[dc_number] = {
                    'dc_number': row[0],
                    'dc_type': row[1],
                    'party_name': row[2],
                    'status':row[3],
                    'item_names': [],
                    'created_date': row[5],
                    'created_by': row[6]
                    
                }
            dc_map[dc_number]['item_names'].append(row[4])

        overalldc_list = list(dc_map.values())

        return render(request, 'overall_dc.html', {'overalldc': overalldc_list})

    except Exception as e:
        logger.error(f"Error fetching overall DC list: {str(e)}", exc_info=True)
        return render(request, 'overall_dc.html', {'overalldc': [], 'error': str(e)})
    
    
    
    
    
    
    
    
    
# vfdndfgbfdgfd/

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.db import connection, DatabaseError, transaction
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from functools import wraps
import logging
import csv
from io import StringIO
import openpyxl
from datetime import datetime

# Set up logging
logger = logging.getLogger(__name__)

# Custom decorator to check if user is admin
def admin_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT role FROM users WHERE username = %s AND role = 'admin'",
                    [request.session.get('username')]
                )
                if not cursor.fetchone():
                    if request.is_ajax():
                        return JsonResponse({'error': 'Unauthorized: Admin access required'}, status=401)
                    return HttpResponse('Unauthorized: Admin access required', status=401)
        except DatabaseError as e:
            logger.error(f"Database error in admin_required: {str(e)}")
            return JsonResponse({'error': 'Database error during authentication'}, status=500)
        return view_func(request, *args, **kwargs)
    return wrapper

@login_required
@admin_required
def dept_wise_dc_view(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT dc_type FROM dc_types WHERE dc_type IS NOT NULL")
            departments = [row[0] for row in cursor.fetchall()]
        return render(request, 'dept_wise_dc.html', {'departments': departments})
    except DatabaseError as e:
        logger.error(f"Error fetching departments: {str(e)}")
        return HttpResponse('Error fetching departments', status=500)

@login_required
@admin_required
@csrf_exempt
def dept_wise_dc_api(request):
    department = request.GET.get('department', '')
    try:
        with connection.cursor() as cursor:
            # Fetch DC list
            if department:
                cursor.execute("""
                    SELECT dd.id, dd.dc_number, dd.dc_date, pd.party_name, dd.status, dd.created_by
                    FROM dc_details dd
                    LEFT JOIN party_details pd ON dd.party_name = pd.party_name
                    WHERE dd.dc_type = %s
                """, [department])
            else:
                cursor.execute("""
                    SELECT dd.id, dd.dc_number, dd.dc_date, pd.party_name, dd.status, dd.created_by
                    FROM dc_details dd
                    LEFT JOIN party_details pd ON dd.party_name = pd.party_name
                """)
            dc_list = [
                {
                    'id': row[0],
                    'dc_number': row[1],
                    'date': row[2].strftime('%Y-%m-%d') if row[2] else '',
                    'party': row[3] if row[3] is not None else 'N/A',
                    'status': row[4] if row[4] is not None else 'N/A',
                    'created_by': row[5] if row[5] is not None else 'N/A'
                }
                for row in cursor.fetchall()
            ]

            # Summary statistics
            if department:
                cursor.execute("""
                    SELECT 
                        COUNT(*) as total,
                        SUM(CASE WHEN status = 'PENDING' THEN 1 ELSE 0 END) as pending,
                        SUM(CASE WHEN status = 'CLOSED' THEN 1 ELSE 0 END) as closed,
                        SUM(CASE WHEN status = 'PARTIAL' THEN 1 ELSE 0 END) as partial
                    FROM dc_details
                    WHERE dc_type = %s
                """, [department])
            else:
                cursor.execute("""
                    SELECT 
                        COUNT(*) as total,
                        SUM(CASE WHEN status = 'PENDING' THEN 1 ELSE 0 END) as pending,
                        SUM(CASE WHEN status = 'CLOSED' THEN 1 ELSE 0 END) as closed,
                        SUM(CASE WHEN status = 'PARTIAL' THEN 1 ELSE 0 END) as partial
                    FROM dc_details
                """)
            summary_row = cursor.fetchone()
            summary = {
                'total': summary_row[0] if summary_row else 0,
                'open': summary_row[1] if summary_row else 0,
                'closed': summary_row[2] if summary_row else 0,
                'partial': summary_row[3] if summary_row else 0
            }

            # Department-wise summary for pie chart
            dept_summary = []
            if not department:
                cursor.execute("""
                    SELECT dc_type, COUNT(*) as count
                    FROM dc_details
                    WHERE dc_type IS NOT NULL AND dc_type != ''
                    GROUP BY dc_type
                """)
                dept_summary = [
                    {'department': row[0], 'count': row[1]}
                    for row in cursor.fetchall()
                ]

            # Monthly trend data for line chart
            if department:
                cursor.execute("""
                    SELECT DATE_FORMAT(dc_date, '%%b %%Y') as month, COUNT(*) as count
                    FROM dc_details
                    WHERE dc_type = %s
                    GROUP BY DATE_FORMAT(dc_date, '%%b %%Y')
                    ORDER BY MIN(dc_date)
                """, [department])
            else:
                cursor.execute("""
                    SELECT DATE_FORMAT(dc_date, '%%b %%Y') as month, COUNT(*) as count
                    FROM dc_details
                    GROUP BY DATE_FORMAT(dc_date, '%%b %%Y')
                    ORDER BY MIN(dc_date)
                """)
            monthly_trend = [
                {'month': row[0], 'count': row[1]}
                for row in cursor.fetchall()
            ]

        return JsonResponse({
            'dcs': dc_list,
            'summary': summary,
            'dept_summary': dept_summary,
            'monthly_trend': monthly_trend
        })
    except DatabaseError as e:
        logger.error(f"Error in dept_wise_dc_api: {str(e)}")
        return JsonResponse({'error': f'Database error: {str(e)}'}, status=500)
    except Exception as e:
        logger.error(f"Unexpected error in dept_wise_dc_api: {str(e)}")
        return JsonResponse({'error': f'Unexpected error: {str(e)}'}, status=500)

@login_required
@admin_required
def dept_wise_report(request):
    department = request.GET.get('department', '')
    try:
        with connection.cursor() as cursor:
            if department:
                cursor.execute("""
                    SELECT dd.dc_number, dd.dc_date, pd.party_name, dd.status, dd.created_by
                    FROM dc_details dd
                    LEFT JOIN party_details pd ON dd.party_name = pd.party_name
                    WHERE dd.dc_type = %s
                """, [department])
            else:
                cursor.execute("""
                    SELECT dd.dc_number, dd.dc_date, pd.party_name, dd.status, dd.created_by
                    FROM dc_details dd
                    LEFT JOIN party_details pd ON dd.party_name = pd.party_name
                """)
            rows = cursor.fetchall()
            logger.debug(f"Query returned {len(rows)} rows for department: {department}")
            if not rows:
                logger.warning(f"No data found for department: {department}")
                # Return empty CSV with headers
                output = StringIO()
                writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
                writer.writerow(['DC Number', 'Date', 'Party', 'Status', 'Created By'])
                response = HttpResponse(
                    content_type='text/csv',
                    headers={'Content-Disposition': f'attachment; filename="dc_report_{department or "all"}.csv"'}
                )
                response.write(output.getvalue())
                output.close()
                return response

            report_data = [
                {
                    'dc_number': row[0],
                    'date': row[1].strftime('%Y-%m-%d') if row[1] else '',
                    'party': row[2] if row[2] is not None else 'N/A',
                    'status': row[3] if row[3] is not None else 'N/A',
                    'created_by': row[4] if row[4] is not None else 'N/A'
                }
                for row in rows
            ]

        # Generate CSV
        output = StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(['DC Number', 'Date', 'Party', 'Status', 'Created By'])
        for row in report_data:
            writer.writerow([
                row['dc_number'],
                row['date'],
                row['party'],
                row['status'],
                row['created_by']
            ])

        response = HttpResponse(
            content_type='text/csv',
            headers={'Content-Disposition': f'attachment; filename="dc_report_{department or "all"}.csv"'}
        )
        response.write(output.getvalue())
        output.close()
        return response
    except DatabaseError as e:
        logger.error(f"Database error in dept_wise_report: {str(e)}")
        return HttpResponse(f'Database error: {str(e)}', status=500)
    except IndexError as e:
        logger.error(f"IndexError in dept_wise_report: {str(e)}, department: {department}, rows: {rows}")
        return HttpResponse(f'IndexError: Invalid data format in database query', status=500)
    except Exception as e:
        logger.error(f"Unexpected error in dept_wise_report: {str(e)}")
        return HttpResponse(f'Unexpected error: {str(e)}', status=500)

@login_required
@admin_required
def dc_details_report(request, dc_number):
    try:
        with connection.cursor() as cursor:
            # Fetch dc_details
            cursor.execute("""
                SELECT dd.dc_number, dd.dc_date, pd.party_name, dd.status, dd.created_by,
                       dd.vehicle_no, dd.process, dd.total_dispatch_quantity, dd.dc_type,
                       dd.created_date_time, dd.updated_date_time
                FROM dc_details dd
                LEFT JOIN party_details pd ON dd.party_name = pd.party_name
                WHERE dd.dc_number = %s
            """, [dc_number])
            dc_details_row = cursor.fetchone()
            logger.debug(f"dc_details query for {dc_number}: {dc_details_row}")
            if not dc_details_row:
                logger.warning(f"No data found in dc_details for dc_number: {dc_number}")
                return HttpResponse(f'No data found for DC number: {dc_number}', status=404)

            dc_details = {
                'dc_number': dc_details_row[0],
                'date': dc_details_row[1].strftime('%Y-%m-%d') if dc_details_row[1] else 'N/A',
                'party': dc_details_row[2] if dc_details_row[2] is not None else 'N/A',
                'status': dc_details_row[3] if dc_details_row[3] is not None else 'N/A',
                'created_by': dc_details_row[4] if dc_details_row[4] is not None else 'N/A',
                'vehicle_no': dc_details_row[5] if dc_details_row[5] is not None else 'N/A',
                'process': dc_details_row[6] if dc_details_row[6] is not None else 'N/A',
                'total_dispatch_quantity': dc_details_row[7] if dc_details_row[7] is not None else '0.00',
                'dc_type': dc_details_row[8] if dc_details_row[8] is not None else 'N/A',
                'created_date_time': dc_details_row[9].strftime('%Y-%m-%d %H:%M:%S') if dc_details_row[9] else 'N/A',
                'updated_date_time': dc_details_row[10].strftime('%Y-%m-%d %H:%M:%S') if dc_details_row[10] else 'N/A'
            }

            # Fetch dc_description
            cursor.execute("""
                SELECT id, dc_number, item_name, quantity
                FROM dc_description
                WHERE dc_number = %s
            """, [dc_number])
            dc_description_rows = cursor.fetchall()
            logger.debug(f"dc_description query for {dc_number}: {len(dc_description_rows)} rows")
            dc_description = [
                {
                    'id': row[0],
                    'dc_number': row[1],
                    'item_name': row[2] if row[2] is not None else 'N/A',
                    'quantity': row[3] if row[3] is not None else '0.00'
                }
                for row in dc_description_rows
            ]

            # Fetch party_details
            cursor.execute("""
                SELECT party_name, address, state
                FROM party_details
                WHERE party_name = %s
            """, [dc_details['party']])
            party_details_row = cursor.fetchone()
            logger.debug(f"party_details query for {dc_details['party']}: {party_details_row}")
            party_details = {
                'party_name': party_details_row[0] if party_details_row and party_details_row[0] is not None else 'N/A',
                'address': party_details_row[1] if party_details_row and party_details_row[1] is not None else 'N/A',
                'state': party_details_row[2] if party_details_row and party_details_row[2] is not None else 'N/A'
            } if party_details_row else {'party_name': 'N/A', 'address': 'N/A', 'state': 'N/A'}

            # Fetch party_dc
            cursor.execute("""
                SELECT party_dc_number, dc_number, party_dc_date, party_dc_doc_name
                FROM party_dc
                WHERE dc_number = %s
            """, [dc_number])
            party_dc_rows = cursor.fetchall()
            logger.debug(f"party_dc query for {dc_number}: {len(party_dc_rows)} rows")
            party_dc = [
                {
                    'party_dc_number': row[0] if row[0] is not None else 'N/A',
                    'dc_number': row[1] if row[1] is not None else 'N/A',
                    'party_dc_date': row[2].strftime('%Y-%m-%d') if row[2] else 'N/A',
                    'party_dc_doc_name': row[3] if row[3] is not None else 'N/A'
                }
                for row in party_dc_rows
            ]

            # Fetch party_dc_details
            cursor.execute("""
                SELECT pd.party_dc_number, pd.item_name, pd.received_quantity
                FROM party_dc_details pd
                JOIN party_dc p ON pd.party_dc_number = p.party_dc_number
                WHERE p.dc_number = %s
            """, [dc_number])
            party_dc_details_rows = cursor.fetchall()
            logger.debug(f"party_dc_details query for {dc_number}: {len(party_dc_details_rows)} rows")
            party_dc_details = [
                {
                    'party_dc_number': row[0] if row[0] is not None else 'N/A',
                    'item_name': row[1] if row[1] is not None else 'N/A',
                    'received_quantity': row[2] if row[2] is not None else '0.00'
                }
                for row in party_dc_details_rows
            ]

        # Generate Excel
        wb = openpyxl.Workbook()

        # dc_details sheet
        ws_dc = wb.active
        ws_dc.title = "DC Details"
        headers_dc = ['DC Number', 'Date', 'Party', 'Status', 'Created By', 'Vehicle No', 'Process', 'Total Dispatch Quantity', 'Department', 'Created Date Time', 'Updated Date Time']
        ws_dc.append(headers_dc)
        ws_dc.append([
            dc_details['dc_number'],
            dc_details['date'],
            dc_details['party'],
            dc_details['status'],
            dc_details['created_by'],
            dc_details['vehicle_no'],
            dc_details['process'],
            dc_details['total_dispatch_quantity'],
            dc_details['dc_type'],
            dc_details['created_date_time'],
            dc_details['updated_date_time']
        ])

        # dc_description sheet
        ws_desc = wb.create_sheet("DC Description")
        headers_desc = ['ID', 'DC Number', 'Item Name', 'Quantity']
        ws_desc.append(headers_desc)
        for row in dc_description:
            ws_desc.append([
                row['id'],
                row['dc_number'],
                row['item_name'],
                row['quantity']
            ])

        # party_details sheet
        ws_party = wb.create_sheet("Party Details")
        headers_party = ['Party Name', 'Address', 'State']
        ws_party.append(headers_party)
        ws_party.append([
            party_details['party_name'],
            party_details['address'],
            party_details['state']
        ])

        # party_dc sheet
        ws_party_dc = wb.create_sheet("Party DC")
        headers_party_dc = ['Party DC Number', 'DC Number', 'Party DC Date', 'Party DC Doc Name']
        ws_party_dc.append(headers_party_dc)
        for row in party_dc:
            ws_party_dc.append([
                row['party_dc_number'],
                row['dc_number'],
                row['party_dc_date'],
                row['party_dc_doc_name']
            ])

        # party_dc_details sheet
        ws_party_dc_details = wb.create_sheet("Party DC Details")
        headers_party_dc_details = ['Party DC Number', 'Item Name', 'Received Quantity']
        ws_party_dc_details.append(headers_party_dc_details)
        for row in party_dc_details:
            ws_party_dc_details.append([
                row['party_dc_number'],
                row['item_name'],
                row['received_quantity']
            ])

        # Adjust column widths for all sheets
        for ws in [ws_dc, ws_desc, ws_party, ws_party_dc, ws_party_dc_details]:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="dc_details_{dc_number}.xlsx"'}
        )
        wb.save(response)
        return response
    except DatabaseError as e:
        logger.error(f"Database error in dc_details_report for dc_number {dc_number}: {str(e)}")
        return HttpResponse(f'Database error: {str(e)}', status=500)
    except Exception as e:
        logger.error(f"Unexpected error in dc_details_report for dc_number {dc_number}: {str(e)}")
        return HttpResponse(f'Unexpected error: {str(e)}', status=500)

@login_required
@admin_required
@csrf_exempt
def delete_dc(request, dc_id):
    try:
        with connection.cursor() as cursor:
            # Start a transaction to ensure atomicity
            with transaction.atomic():
                # 1. Fetch and copy dc_details to deleted_dc_details
                cursor.execute("""
                    SELECT id, dc_number, dc_date, party_name, status, created_by,
                           vehicle_no, process, total_dispatch_quantity, dc_type,
                           created_date_time, updated_date_time
                    FROM dc_details
                    WHERE id = %s
                """, [dc_id])
                dc_details_row = cursor.fetchone()
                if not dc_details_row:
                    logger.warning(f"No data found in dc_details for id: {dc_id}")
                    return JsonResponse({'error': 'DC not found'}, status=404)

                cursor.execute("""
                    INSERT INTO deleted_dc_details (
                        id, dc_number, dc_date, party_name, status, created_by,
                        vehicle_no, process, total_dispatch_quantity, dc_type,
                        created_date_time, updated_date_time, deleted_at
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                """, list(dc_details_row))

                # 2. Fetch and copy dc_description to deleted_dc_description
                cursor.execute("""
                    SELECT id, dc_number, item_name, quantity
                    FROM dc_description
                    WHERE dc_number = (SELECT dc_number FROM dc_details WHERE id = %s)
                """, [dc_id])
                dc_description_rows = cursor.fetchall()
                for row in dc_description_rows:
                    cursor.execute("""
                        INSERT INTO deleted_dc_description (
                            id, dc_number, item_name, quantity, deleted_at
                        )
                        VALUES (%s, %s, %s, %s, NOW())
                    """, list(row))

                # 3. Fetch and copy party_dc to deleted_party_dc
                cursor.execute("""
                    SELECT id, party_dc_number, dc_number, party_dc_date, party_dc_doc_name
                    FROM party_dc_details
                    WHERE dc_number = (SELECT dc_number FROM dc_details WHERE id = %s)
                """, [dc_id])
                party_dc_rows = cursor.fetchall()
                for row in party_dc_rows:
                    cursor.execute("""
                        INSERT INTO deleted_party_dc (
                            id, party_dc_number, dc_number, party_dc_date, party_dc_doc_name, deleted_at
                        )
                        VALUES (%s, %s, %s, %s, %s, NOW())
                    """, list(row))

               

                # 5. Delete records from original tables
                cursor.execute("""
                    DELETE FROM party_dc_details
                    WHERE party_dc_number IN (
                        SELECT party_dc_number FROM party_dc WHERE dc_number = (
                            SELECT dc_number FROM dc_details WHERE id = %s
                        )
                    )
                """, [dc_id])
                cursor.execute("""
                    DELETE FROM party_dc
                    WHERE dc_number = (SELECT dc_number FROM dc_details WHERE id = %s)
                """, [dc_id])
                cursor.execute("""
                    DELETE FROM dc_description
                    WHERE dc_number = (SELECT dc_number FROM dc_details WHERE id = %s)
                """, [dc_id])
                cursor.execute("DELETE FROM dc_details WHERE id = %s", [dc_id])

                if cursor.rowcount == 0:
                    logger.warning(f"No records deleted from dc_details for id: {dc_id}")
                    return JsonResponse({'error': 'DC not found'}, status=404)

        return JsonResponse({'status': 'success'})
    except DatabaseError as e:
        logger.error(f"Database error in delete_dc for id {dc_id}: {str(e)}")
        return JsonResponse({'error': f'Database error: {str(e)}'}, status=500)
    except Exception as e:
        logger.error(f"Unexpected error in delete_dc for id {dc_id}: {str(e)}")
        return JsonResponse({'error': f'Unexpected error: {str(e)}'}, status=500)

def cancelled_dc_list(request):
    """List delivery challans with status CANCELLED for the authenticated user"""
    username = request.session.get('username', '')
    if not username:
        return redirect('/')
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT dc.id, dc.dc_number, dc.party_name, dc.dc_date, dc.dc_type
                FROM dc_details dc
                WHERE dc.created_by = %s AND dc.status = 'CANCELLED'
                ORDER BY dc.dc_date DESC
            """, [username])
            delivery_challans = cursor.fetchall()
            dc_list = []
            for dc in delivery_challans:
                dc_id, dc_number, party_name, dc_date, dc_type = dc
                cursor.execute("""
                    SELECT item_name
                    FROM dc_description
                    WHERE dc_number = %s
                """, [dc_number])
                items = [row[0] for row in cursor.fetchall() if row[0]]
                # Ensure all fields have default values if they're None
                dc_list.append({
                    'id': dc_id or 0,
                    'dcNo': dc_number or 'N/A',
                    'buyerName': party_name or 'N/A',
                    'dcDate': dc_date or None,
                    'dc_type': dc_type or 'N/A',
                    'items': items if items else ['N/A']
                })
        return render(request, 'cancelled_dc.html', {
            'username': username,
            'delivery_challans': dc_list
        })
    except Exception as e:
        logger.error(f"Error fetching cancelled delivery challan list: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)


def deleted_dc_list(request):
    """List delivery challans from deleted_dc_details table for the authenticated user"""
    username = request.session.get('username', '')
    if not username:
        return redirect('/')
    try:
        with connection.cursor() as cursor:
            # First, let's check if dc_type column exists in deleted_dc_details
            cursor.execute("""
                SELECT COLUMN_NAME 
                FROM INFORMATION_SCHEMA.COLUMNS 
                WHERE TABLE_NAME = 'deleted_dc_details' 
                AND COLUMN_NAME = 'dc_type'
            """)
            has_dc_type = cursor.fetchone() is not None
            
            if has_dc_type:
                cursor.execute("""
                    SELECT dc.id, dc.dc_number, dc.party_name, dc.dc_date, dc.dc_type
                    FROM deleted_dc_details dc
                    WHERE dc.created_by = %s
                    ORDER BY dc.dc_date DESC
                """, [username])
            else:
                cursor.execute("""
                    SELECT dc.id, dc.dc_number, dc.party_name, dc.dc_date
                    FROM deleted_dc_details dc
                    WHERE dc.created_by = %s
                    ORDER BY dc.dc_date DESC
                """, [username])
            
            delivery_challans = cursor.fetchall()
            dc_list = []
            for dc in delivery_challans:
                if has_dc_type:
                    dc_id, dc_number, party_name, dc_date, dc_type = dc
                else:
                    dc_id, dc_number, party_name, dc_date = dc
                    dc_type = 'N/A'  # Default value if dc_type doesn't exist
                
                # Check if deleted_dc_description table exists
                cursor.execute("""
                    SELECT COUNT(*) 
                    FROM INFORMATION_SCHEMA.TABLES 
                    WHERE TABLE_NAME = 'deleted_dc_description'
                """)
                table_exists = cursor.fetchone()[0] > 0
                
                if table_exists:
                    cursor.execute("""
                        SELECT item_name
                        FROM deleted_dc_description
                        WHERE dc_number = %s
                    """, [dc_number])
                    items = [row[0] for row in cursor.fetchall() if row[0]]
                else:
                    # If table doesn't exist, try to get items from original dc_description
                    cursor.execute("""
                        SELECT item_name
                        FROM dc_description
                        WHERE dc_number = %s
                    """, [dc_number])
                    items = [row[0] for row in cursor.fetchall() if row[0]]
                dc_list.append({
                    'id': dc_id,
                    'dcNo': dc_number,
                    'buyerName': party_name,
                    'dcDate': dc_date,
                    'dc_type': dc_type,
                    'items': items
                })
        return render(request, 'deleted_dc.html', {
            'username': username,
            'delivery_challans': dc_list
        })
    except Exception as e:
        logger.error(f"Error fetching deleted delivery challan list: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)


@login_required
@admin_required
@csrf_exempt
def cancel_dc(request, dc_id):
    """Cancel a delivery challan by updating its status to CANCELLED"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    try:
        with connection.cursor() as cursor:
            # Check if DC exists and get current status
            cursor.execute("""
                SELECT status, dc_number FROM dc_details 
                WHERE id = %s
            """, [dc_id])
            result = cursor.fetchone()
            
            if not result:
                return JsonResponse({'error': 'DC not found'}, status=404)
            
            current_status, dc_number = result
            
            # Only allow cancellation if DC is not already cancelled or deleted
            if current_status in ['CANCELLED', 'DELETED']:
                return JsonResponse({'error': f'DC is already {current_status.lower()}'}, status=400)
            
            # Update status to CANCELLED
            cursor.execute("""
                UPDATE dc_details 
                SET status = 'CANCELLED', updated_date_time = NOW()
                WHERE id = %s
            """, [dc_id])
            
            if cursor.rowcount == 0:
                return JsonResponse({'error': 'Failed to cancel DC'}, status=500)
            
            logger.info(f"DC {dc_number} (ID: {dc_id}) has been cancelled")
            return JsonResponse({'status': 'success', 'message': 'DC cancelled successfully'})
            
    except Exception as e:
        logger.error(f"Error cancelling DC {dc_id}: {str(e)}")
        return JsonResponse({'error': f'Error cancelling DC: {str(e)}'}, status=500)


def deleted_dc_view(request, dc_id):
    """View details of a deleted delivery challan"""
    username = request.session.get('username', '')
    if not username:
        return redirect('/')
    
    try:
        with connection.cursor() as cursor:
            # Get DC details from deleted_dc_details
            cursor.execute("""
                SELECT id, dc_number, dc_date, party_name, status, created_by,
                       vehicle_no, process, total_dispatch_quantity, dc_type,
                       created_date_time, updated_date_time
                FROM deleted_dc_details
                WHERE id = %s
            """, [dc_id])
            dc_details_row = cursor.fetchone()
            
            if not dc_details_row:
                return HttpResponse("DC not found", status=404)
            
            # Get party details
            cursor.execute("""
                SELECT party_name, address_line_1, address_line_2, city, state, pincode, gstin_number
                FROM party_details
                WHERE party_name = %s
            """, [dc_details_row[3]])  # party_name
            party_details = cursor.fetchone()
            
            # Get DC description items
            cursor.execute("""
                SELECT item_name, quantity
                FROM deleted_dc_description
                WHERE dc_number = %s
            """, [dc_details_row[1]])  # dc_number
            dc_description_rows = cursor.fetchall()
            
            # Get party DC details
            cursor.execute("""
                SELECT party_dc_number, party_dc_date, party_dc_doc_name
                FROM deleted_party_dc
                WHERE dc_number = %s
            """, [dc_details_row[1]])  # dc_number
            party_dc_rows = cursor.fetchall()
            
            # Prepare context data
            dc_details = {
                'id': dc_details_row[0],
                'dc_number': dc_details_row[1],
                'dc_date': dc_details_row[2],
                'party_name': dc_details_row[3],
                'status': dc_details_row[4],
                'created_by': dc_details_row[5],
                'vehicle_no': dc_details_row[6],
                'process': dc_details_row[7],
                'total_dispatch_quantity': dc_details_row[8],
                'dc_type': dc_details_row[9] if len(dc_details_row) > 9 else 'N/A',
                'created_date_time': dc_details_row[10],
                'updated_date_time': dc_details_row[11]
            }
            
            party_data = {
                'party_name': party_details[0] if party_details else 'N/A',
                'address_line_1': party_details[1] if party_details else 'N/A',
                'address_line_2': party_details[2] if party_details else 'N/A',
                'city': party_details[3] if party_details else 'N/A',
                'state': party_details[4] if party_details else 'N/A',
                'pincode': party_details[5] if party_details else 'N/A',
                'gstin_number': party_details[6] if party_details else 'N/A'
            }
            
            dc_description = []
            for row in dc_description_rows:
                dc_description.append({
                    'item_name': row[0],
                    'quantity': row[1]
                })
            
            party_dc_details = []
            for row in party_dc_rows:
                party_dc_details.append({
                    'party_dc_number': row[0],
                    'party_dc_date': row[1],
                    'party_dc_doc_name': row[2]
                })
            
            context = {
                'username': username,
                'dc_details': dc_details,
                'party_data': party_data,
                'dc_description': dc_description,
                'party_dc_details': party_dc_details,
                'is_deleted': True  # Flag to indicate this is a deleted DC view
            }
            
            return render(request, 'deleted_dc_view.html', context)
            
    except Exception as e:
        logger.error(f"Error viewing deleted DC {dc_id}: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)


@csrf_exempt
def restore_dc(request, dc_id):
    """Restore a deleted delivery challan by copying data back from deleted tables"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    try:
        with connection.cursor() as cursor:
            # Start a transaction to ensure atomicity
            with transaction.atomic():
                # 1. Check if DC exists in deleted_dc_details
                cursor.execute("""
                    SELECT id, dc_number, dc_date, party_name, status, created_by,
                           vehicle_no, process, total_dispatch_quantity, dc_type,
                           created_date_time, updated_date_time
                    FROM deleted_dc_details
                    WHERE id = %s
                """, [dc_id])
                dc_details_row = cursor.fetchone()
                if not dc_details_row:
                    return JsonResponse({'error': 'DC not found in deleted records'}, status=404)

                # 2. Check if DC number already exists in active dc_details
                cursor.execute("""
                    SELECT id FROM dc_details WHERE dc_number = %s
                """, [dc_details_row[1]])  # dc_number is at index 1
                if cursor.fetchone():
                    return JsonResponse({'error': 'DC number already exists in active records'}, status=400)

                # 3. Restore dc_details
                cursor.execute("""
                    INSERT INTO dc_details (
                        id, dc_number, dc_date, party_name, status, created_by,
                        vehicle_no, process, total_dispatch_quantity, dc_type,
                        created_date_time, updated_date_time
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, list(dc_details_row))

                # 4. Restore dc_description
                cursor.execute("""
                    SELECT id, dc_number, item_name, quantity
                    FROM deleted_dc_description
                    WHERE dc_number = %s
                """, [dc_details_row[1]])  # dc_number
                dc_description_rows = cursor.fetchall()
                for row in dc_description_rows:
                    cursor.execute("""
                        INSERT INTO dc_description (
                            id, dc_number, item_name, quantity
                        )
                        VALUES (%s, %s, %s, %s)
                    """, list(row))

                # 5. Restore party_dc_details
                cursor.execute("""
                    SELECT id, party_dc_number, dc_number, party_dc_date, party_dc_doc_name
                    FROM deleted_party_dc
                    WHERE dc_number = %s
                """, [dc_details_row[1]])  # dc_number
                party_dc_rows = cursor.fetchall()
                for row in party_dc_rows:
                    cursor.execute("""
                        INSERT INTO party_dc_details (
                            id, party_dc_number, dc_number, party_dc_date, party_dc_doc_name
                        )
                        VALUES (%s, %s, %s, %s, %s)
                    """, list(row))

                # 6. Delete from deleted tables
                cursor.execute("DELETE FROM deleted_party_dc WHERE dc_number = %s", [dc_details_row[1]])
                cursor.execute("DELETE FROM deleted_dc_description WHERE dc_number = %s", [dc_details_row[1]])
                cursor.execute("DELETE FROM deleted_dc_details WHERE id = %s", [dc_id])

                logger.info(f"DC {dc_details_row[1]} (ID: {dc_id}) has been restored")
                return JsonResponse({'status': 'success', 'message': 'DC restored successfully'})
                
    except Exception as e:
        logger.error(f"Error restoring DC {dc_id}: {str(e)}")
        return JsonResponse({'error': f'Error restoring DC: {str(e)}'}, status=500)


@login_required
@admin_required
@csrf_exempt
def delete_dc(request, dc_id):
    try:
        with connection.cursor() as cursor:
            # Start a transaction to ensure atomicity
            with transaction.atomic():
                # 1. Fetch and copy dc_details to deleted_dc_details
                cursor.execute("""
                    SELECT id, dc_number, dc_date, party_name, status, created_by,
                           vehicle_no, process, total_dispatch_quantity, dc_type,
                           created_date_time, updated_date_time
                    FROM dc_details
                    WHERE id = %s
                """, [dc_id])
                dc_details_row = cursor.fetchone()
                if not dc_details_row:
                    logger.warning(f"No data found in dc_details for id: {dc_id}")
                    return JsonResponse({'error': 'DC not found'}, status=404)

                cursor.execute("""
                    INSERT INTO deleted_dc_details (
                        id, dc_number, dc_date, party_name, status, created_by,
                        vehicle_no, process, total_dispatch_quantity, dc_type,
                        created_date_time, updated_date_time, deleted_at
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                """, list(dc_details_row))

                # 2. Fetch and copy dc_description to deleted_dc_description
                cursor.execute("""
                    SELECT id, dc_number, item_name, quantity
                    FROM dc_description
                    WHERE dc_number = (SELECT dc_number FROM dc_details WHERE id = %s)
                """, [dc_id])
                dc_description_rows = cursor.fetchall()
                for row in dc_description_rows:
                    cursor.execute("""
                        INSERT INTO deleted_dc_description (
                            id, dc_number, item_name, quantity, deleted_at
                        )
                        VALUES (%s, %s, %s, %s, NOW())
                    """, list(row))

                # 3. Fetch and copy party_dc to deleted_party_dc
                cursor.execute("""
                    SELECT id, party_dc_number, dc_number, party_dc_date, party_dc_doc_name
                    FROM party_dc_details
                    WHERE dc_number = (SELECT dc_number FROM dc_details WHERE id = %s)
                """, [dc_id])
                party_dc_rows = cursor.fetchall()
                for row in party_dc_rows:
                    cursor.execute("""
                        INSERT INTO deleted_party_dc (
                            id, party_dc_number, dc_number, party_dc_date, party_dc_doc_name, deleted_at
                        )
                        VALUES (%s, %s, %s, %s, %s, NOW())
                    """, list(row))

               

                # 5. Delete records from original tables
                cursor.execute("""
                    DELETE FROM party_dc_details
                    WHERE party_dc_number IN (
                        SELECT party_dc_number FROM party_dc WHERE dc_number = (
                            SELECT dc_number FROM dc_details WHERE id = %s
                        )
                    )
                """, [dc_id])
                cursor.execute("""
                    DELETE FROM party_dc
                    WHERE dc_number = (SELECT dc_number FROM dc_details WHERE id = %s)
                """, [dc_id])
                cursor.execute("""
                    DELETE FROM dc_description
                    WHERE dc_number = (SELECT dc_number FROM dc_details WHERE id = %s)
                """, [dc_id])
                cursor.execute("DELETE FROM dc_details WHERE id = %s", [dc_id])

                if cursor.rowcount == 0:
                    logger.warning(f"No records deleted from dc_details for id: {dc_id}")
                    return JsonResponse({'error': 'DC not found'}, status=404)

        return JsonResponse({'status': 'success'})
    except DatabaseError as e:
        logger.error(f"Database error in delete_dc for id {dc_id}: {str(e)}")
        return JsonResponse({'error': f'Database error: {str(e)}'}, status=500)
    except Exception as e:
        logger.error(f"Unexpected error in delete_dc for id {dc_id}: {str(e)}")
        return JsonResponse({'error': f'Unexpected error: {str(e)}'}, status=500)

@login_required
def party_wise_dc_list(request):
    """Show DCs grouped by party with summary information."""
    try:
        with connection.cursor() as cursor:
            # Get all parties with their DC counts and summary
            cursor.execute("""
                SELECT 
                    p.party_name,
                    p.city,
                    p.state,
                    COUNT(dc.id) as total_dcs,
                    SUM(CASE WHEN dc.status = 'PENDING' THEN 1 ELSE 0 END) as pending_dcs,
                    SUM(CASE WHEN dc.status = 'PARTIAL' THEN 1 ELSE 0 END) as partial_dcs,
                    SUM(CASE WHEN dc.status = 'CLOSED' THEN 1 ELSE 0 END) as closed_dcs,
                    SUM(CASE WHEN dc.status = 'CANCELLED' THEN 1 ELSE 0 END) as cancelled_dcs,
                    MAX(dc.dc_date) as latest_dc_date,
                    MIN(dc.dc_date) as earliest_dc_date
                FROM party_details p
                LEFT JOIN dc_details dc ON p.party_name = dc.party_name
                GROUP BY p.party_name, p.city, p.state
                HAVING total_dcs > 0
                ORDER BY total_dcs DESC, p.party_name
            """)
            
            parties = []
            for row in cursor.fetchall():
                parties.append({
                    'party_name': row[0],
                    'city': row[1],
                    'state': row[2],
                    'total_dcs': row[3],
                    'pending_dcs': row[4],
                    'partial_dcs': row[5],
                    'closed_dcs': row[6],
                    'cancelled_dcs': row[7],
                    'latest_dc_date': row[8],
                    'earliest_dc_date': row[9]
                })
        
        return render(request, 'party_wise_dc.html', {'parties': parties})
    except Exception as e:
        logger.error(f"Error in party_wise_dc_list: {str(e)}")
        return render(request, 'party_wise_dc.html', {'parties': [], 'error': str(e)})


@login_required
def party_wise_dc_details(request, party_name):
    """Show all DCs for a specific party."""
    try:
        with connection.cursor() as cursor:
            # Get party details
            cursor.execute("""
                SELECT party_name, address_line_1, address_line_2, city, state, pincode, gstin_number
                FROM party_details
                WHERE party_name = %s
            """, [party_name])
            party_details = cursor.fetchone()
            
            if not party_details:
                return render(request, 'party_wise_dc_details.html', {
                    'error': 'Party not found',
                    'party_name': party_name
                })
            
            # Check if party has any DCs
            cursor.execute("""
                SELECT COUNT(*) FROM dc_details WHERE party_name = %s
            """, [party_name])
            dc_count = cursor.fetchone()[0]
            
            if dc_count == 0:
                return render(request, 'party_wise_dc_details.html', {
                    'error': 'No DCs found for this party',
                    'party_name': party_name,
                    'party_data': {
                        'party_name': party_details[0],
                        'address_line_1': party_details[1],
                        'address_line_2': party_details[2],
                        'city': party_details[3],
                        'state': party_details[4],
                        'pincode': party_details[5],
                        'gstin_number': party_details[6]
                    },
                    'dc_list': [],
                    'summary_data': {
                        'total_dcs': 0,
                        'pending_count': 0,
                        'partial_count': 0,
                        'closed_count': 0,
                        'cancelled_count': 0,
                        'total_dispatch': 0,
                        'total_received': 0,
                        'total_pending': 0,
                        'total_amount': 0
                    }
                })
            
            party_data = {
                'party_name': party_details[0],
                'address_line_1': party_details[1],
                'address_line_2': party_details[2],
                'city': party_details[3],
                'state': party_details[4],
                'pincode': party_details[5],
                'gstin_number': party_details[6]
            }
            
            # Get all DCs for this party
            cursor.execute("""
                SELECT 
                    dc.id,
                    dc.dc_number,
                    dc.dc_date,
                    dc.vehicle_no,
                    dc.process,
                    dc.total_dispatch_quantity,
                    dc.total_received_quantity,
                    dc.total_pending_quantity,
                    dc.total_rate,
                    dc.project_name,
                    dc.project_incharge,
                    dc.status,
                    dc.created_date_time
                FROM dc_details dc
                WHERE dc.party_name = %s
                ORDER BY dc.dc_date DESC
            """, [party_name])
            
            dc_list = []
            for row in cursor.fetchall():
                dc_list.append({
                    'id': row[0],
                    'dc_number': row[1],
                    'dc_date': row[2],
                    'vehicle_no': row[3],
                    'process': row[4],
                    'total_dispatch_quantity': row[5],
                    'total_received_quantity': row[6],
                    'total_pending_quantity': row[7],
                    'total_rate': row[8],
                    'project_name': row[9],
                    'project_incharge': row[10],
                    'status': row[11],
                    'created_date_time': row[12]
                })
            
            # Get summary statistics
            cursor.execute("""
                SELECT 
                    COUNT(*) as total_dcs,
                    SUM(CASE WHEN status = 'PENDING' THEN 1 ELSE 0 END) as pending_count,
                    SUM(CASE WHEN status = 'PARTIAL' THEN 1 ELSE 0 END) as partial_count,
                    SUM(CASE WHEN status = 'CLOSED' THEN 1 ELSE 0 END) as closed_count,
                    SUM(CASE WHEN status = 'CANCELLED' THEN 1 ELSE 0 END) as cancelled_count,
                    SUM(total_dispatch_quantity) as total_dispatch,
                    SUM(total_received_quantity) as total_received,
                    SUM(total_pending_quantity) as total_pending,
                    SUM(total_rate) as total_amount
                FROM dc_details
                WHERE party_name = %s
            """, [party_name])
            
            summary = cursor.fetchone()
            summary_data = {
                'total_dcs': summary[0] if summary and summary[0] else 0,
                'pending_count': summary[1] if summary and summary[1] else 0,
                'partial_count': summary[2] if summary and summary[2] else 0,
                'closed_count': summary[3] if summary and summary[3] else 0,
                'cancelled_count': summary[4] if summary and summary[4] else 0,
                'total_dispatch': summary[5] if summary and summary[5] else 0,
                'total_received': summary[6] if summary and summary[6] else 0,
                'total_pending': summary[7] if summary and summary[7] else 0,
                'total_amount': summary[8] if summary and summary[8] else 0
            }
        
        return render(request, 'party_wise_dc_details.html', {
            'party_data': party_data,
            'dc_list': dc_list,
            'summary_data': summary_data
        })
    except Exception as e:
        logger.error(f"Error in party_wise_dc_details: {str(e)}")
        return render(request, 'party_wise_dc_details.html', {
            'error': str(e),
            'party_name': party_name
        })

def cancelled_dc_view(request, dc_id):
    """Display details of a cancelled delivery challan"""
    username = request.session.get('username', '')
    if not username:
        return redirect('/')
    
    try:
        with connection.cursor() as cursor:
            # Get DC details
            cursor.execute("""
                SELECT dc.id, dc.dc_number, dc.party_name, dc.dc_date, dc.dc_type, 
                       dc.status, dc.created_by, dc.created_date
                FROM dc_details dc
                WHERE dc.id = %s AND dc.status = 'CANCELLED' AND dc.created_by = %s
            """, [dc_id, username])
            dc_data = cursor.fetchone()
            
            if not dc_data:
                return HttpResponse("DC not found or access denied", status=404)
            
            dc_id, dc_number, party_name, dc_date, dc_type, status, created_by, created_date = dc_data
            
            # Get party details
            cursor.execute("""
                SELECT party_name, address_line_1, address_line_2, city, state, pincode, 
                       phone, email, gst_number
                FROM party_details
                WHERE party_name = %s
            """, [party_name])
            party_data = cursor.fetchone()
            
            if party_data:
                party_name, address_line_1, address_line_2, city, state, pincode, phone, email, gst_number = party_data
                party_data = {
                    'name': party_name,
                    'address_line_1': address_line_1,
                    'address_line_2': address_line_2,
                    'city': city,
                    'state': state,
                    'pincode': pincode,
                    'phone': phone,
                    'email': email,
                    'gst_number': gst_number
                }
            else:
                party_data = None
            
            # Get DC description items
            cursor.execute("""
                SELECT item_name, quantity, unit_price, total_price
                FROM dc_description
                WHERE dc_number = %s
                ORDER BY id
            """, [dc_number])
            items_data = cursor.fetchall()
            
            items = []
            total_amount = 0
            for item in items_data:
                item_name, quantity, unit_price, total_price = item
                items.append({
                    'name': item_name,
                    'quantity': quantity,
                    'unit_price': unit_price,
                    'total_price': total_price
                })
                if total_price:
                    total_amount += float(total_price)
            
            dc_info = {
                'id': dc_id,
                'dc_number': dc_number,
                'party_name': party_name,
                'dc_date': dc_date,
                'dc_type': dc_type,
                'status': status,
                'created_by': created_by,
                'created_date': created_date,
                'items': items,
                'total_amount': total_amount
            }
            
        return render(request, 'cancelled_dc_view.html', {
            'username': username,
            'dc': dc_info,
            'party': party_data
        })
        
    except Exception as e:
        logger.error(f"Error fetching cancelled DC details: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)


def restore_cancelled_dc(request, dc_id):
    """Restore a cancelled DC back to PENDING status"""
    username = request.session.get('username', '')
    if not username:
        return JsonResponse({'status': 'error', 'message': 'Not authenticated'}, status=401)
    
    try:
        with connection.cursor() as cursor:
            # Check if DC exists and is cancelled
            cursor.execute("""
                SELECT dc_number, status, created_by
                FROM dc_details
                WHERE id = %s AND status = 'CANCELLED'
            """, [dc_id])
            dc_data = cursor.fetchone()
            
            if not dc_data:
                return JsonResponse({'status': 'error', 'message': 'DC not found or not cancelled'}, status=404)
            
            dc_number, status, created_by = dc_data
            
            # Check if user has permission (admin or creator)
            if created_by != username:
                return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
            
            # Update status to PENDING
            cursor.execute("""
                UPDATE dc_details
                SET status = 'PENDING'
                WHERE id = %s
            """, [dc_id])
            
            logger.info(f"DC {dc_number} (ID: {dc_id}) has been restored from CANCELLED to PENDING by {username}")
            
            return JsonResponse({
                'status': 'success',
                'message': f'DC {dc_number} has been restored successfully'
            })
            
    except Exception as e:
        logger.error(f"Error restoring cancelled DC: {str(e)}")
        return JsonResponse({'status': 'error', 'message': 'Failed to restore DC'}, status=500)

@csrf_exempt
def validate_serial_number(request):
    """Validate serial numbers for valve and actuator"""
    if request.method == 'POST':
        try:
            serial_number = request.POST.get('serial_number', '').strip()
            field_type = request.POST.get('field_type', '').strip()
            
            if not serial_number:
                return JsonResponse({'valid': False, 'message': 'Serial number is required'})
            
            with connection.cursor() as cursor:
                if field_type == 'valve_serial_number':
                    # Check if valve serial number exists
                    cursor.execute(
                        "SELECT id FROM valve_master WHERE serial_number = %s", 
                        [serial_number]
                    )
                elif field_type == 'actuator_serial_number':
                    # Check if actuator serial number exists
                    cursor.execute(
                        "SELECT id FROM actuator_master WHERE serial_number = %s", 
                        [serial_number]
                    )
                else:
                    return JsonResponse({'valid': False, 'message': 'Invalid field type'})
                
                result = cursor.fetchone()
                
                if result:
                    return JsonResponse({'valid': True, 'message': 'Serial number is valid'})
                else:
                    return JsonResponse({'valid': False, 'message': 'Serial number not found in database'})
                    
        except Exception as e:
            logger.error(f"Error validating serial number: {str(e)}")
            return JsonResponse({'valid': False, 'message': 'Error validating serial number'})
    
    return JsonResponse({'valid': False, 'message': 'Invalid request method'})

@csrf_exempt
def get_component_data(request):
    """Get component data based on serial number"""
    if request.method == 'POST':
        try:
            serial_number = request.POST.get('serial_number', '').strip()
            field_type = request.POST.get('field_type', '').strip()
            
            if not serial_number:
                return JsonResponse({'success': False, 'message': 'Serial number is required'})
            
            with connection.cursor() as cursor:
                if field_type == 'valve_serial_number':
                    # Get valve component data
                    cursor.execute("""
                        SELECT 
                            air_shell_status, hydro_shell_status, air_seat_a_status, 
                            air_seat_b_status, valve_status, air_shell_result, 
                            hydro_shell_result, air_seat_a_result, air_seat_b_result, 
                            valve_result, air_shell_remarks, hydro_shell_remarks, 
                            air_seat_a_remarks, air_seat_b_remarks, valve_remarks,
                            cycle_status, completion_date, cycle_remarks
                        FROM valve_test_status 
                        WHERE valve_serial_number = %s
                        ORDER BY created_at DESC LIMIT 1
                    """, [serial_number])
                elif field_type == 'actuator_serial_number':
                    # Get actuator component data
                    cursor.execute("""
                        SELECT 
                            air_shell_status, hydro_shell_status, air_seat_a_status, 
                            air_seat_b_status, valve_status, air_shell_result, 
                            hydro_shell_result, air_seat_a_result, air_seat_b_result, 
                            valve_result, air_shell_remarks, hydro_shell_remarks, 
                            air_seat_a_remarks, air_seat_b_remarks, valve_remarks,
                            cycle_status, completion_date, cycle_remarks
                        FROM valve_test_status 
                        WHERE actuator_serial_number = %s
                        ORDER BY created_at DESC LIMIT 1
                    """, [serial_number])
                else:
                    return JsonResponse({'success': False, 'message': 'Invalid field type'})
                
                result = cursor.fetchone()
                
                if result:
                    data = {
                        'air_shell_status': result[0],
                        'hydro_shell_status': result[1],
                        'air_seat_a_status': result[2],
                        'air_seat_b_status': result[3],
                        'valve_status': result[4],
                        'air_shell_result': result[5],
                        'hydro_shell_result': result[6],
                        'air_seat_a_result': result[7],
                        'air_seat_b_result': result[8],
                        'valve_result': result[9],
                        'air_shell_remarks': result[10],
                        'hydro_shell_remarks': result[11],
                        'air_seat_a_remarks': result[12],
                        'air_seat_b_remarks': result[13],
                        'valve_remarks': result[14],
                        'cycle_status': result[15],
                        'completion_date': result[16],
                        'cycle_remarks': result[17]
                    }
                    return JsonResponse({'success': True, 'data': data})
                else:
                    return JsonResponse({'success': False, 'message': 'No data found for this serial number'})
                    
        except Exception as e:
            logger.error(f"Error getting component data: {str(e)}")
            return JsonResponse({'success': False, 'message': 'Error retrieving component data'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@csrf_exempt
def get_cycle_status_data(request):
    """Get cycle status data for existing serial numbers"""
    if request.method == 'POST':
        try:
            valve_serial = request.POST.get('valve_serial', '').strip()
            actuator_serial = request.POST.get('actuator_serial', '').strip()
            
            if not valve_serial and not actuator_serial:
                return JsonResponse({'success': False, 'message': 'At least one serial number is required'})
            
            with connection.cursor() as cursor:
                if valve_serial and actuator_serial:
                    cursor.execute("""
                        SELECT 
                            air_shell_status, hydro_shell_status, air_seat_a_status, 
                            air_seat_b_status, valve_status, air_shell_result, 
                            hydro_shell_result, air_seat_a_result, air_seat_b_result, 
                            valve_result, air_shell_remarks, hydro_shell_remarks, 
                            air_seat_a_remarks, air_seat_b_remarks, valve_remarks,
                            cycle_status, completion_date, cycle_remarks
                        FROM valve_test_status 
                        WHERE valve_serial_number = %s AND actuator_serial_number = %s
                        ORDER BY created_at DESC LIMIT 1
                    """, [valve_serial, actuator_serial])
                elif valve_serial:
                    cursor.execute("""
                        SELECT 
                            air_shell_status, hydro_shell_status, air_seat_a_status, 
                            air_seat_b_status, valve_status, air_shell_result, 
                            hydro_shell_result, air_seat_a_result, air_seat_b_result, 
                            valve_result, air_shell_remarks, hydro_shell_remarks, 
                            air_seat_a_remarks, air_seat_b_remarks, valve_remarks,
                            cycle_status, completion_date, cycle_remarks
                        FROM valve_test_status 
                        WHERE valve_serial_number = %s
                        ORDER BY created_at DESC LIMIT 1
                    """, [valve_serial])
                else:
                    cursor.execute("""
                        SELECT 
                            air_shell_status, hydro_shell_status, air_seat_a_status, 
                            air_seat_b_status, valve_status, air_shell_result, 
                            hydro_shell_result, air_seat_a_result, air_seat_b_result, 
                            valve_result, air_shell_remarks, hydro_shell_remarks, 
                            air_seat_a_remarks, air_seat_b_remarks, valve_remarks,
                            cycle_status, completion_date, cycle_remarks
                        FROM valve_test_status 
                        WHERE actuator_serial_number = %s
                        ORDER BY created_at DESC LIMIT 1
                    """, [actuator_serial])
                
                result = cursor.fetchone()
                
                if result:
                    data = {
                        'air_shell_status': result[0],
                        'hydro_shell_status': result[1],
                        'air_seat_a_status': result[2],
                        'air_seat_b_status': result[3],
                        'valve_status': result[4],
                        'air_shell_result': result[5],
                        'hydro_shell_result': result[6],
                        'air_seat_a_result': result[7],
                        'air_seat_b_result': result[8],
                        'valve_result': result[9],
                        'air_shell_remarks': result[10],
                        'hydro_shell_remarks': result[11],
                        'air_seat_a_remarks': result[12],
                        'air_seat_b_remarks': result[13],
                        'valve_remarks': result[14],
                        'cycle_status': result[15],
                        'completion_date': result[16],
                        'cycle_remarks': result[17]
                    }
                    return JsonResponse({'success': True, 'data': data})
                else:
                    return JsonResponse({'success': False, 'message': 'No cycle status data found'})
                    
        except Exception as e:
            logger.error(f"Error getting cycle status data: {str(e)}")
            return JsonResponse({'success': False, 'message': 'Error retrieving cycle status data'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@csrf_exempt
def save_cycle_status(request):
    """Save cycle completion status data"""
    if request.method == 'POST':
        try:
            # Get all form data
            valve_serial = request.POST.get('valve_serial_number', '').strip()
            actuator_serial = request.POST.get('actuator_serial_number', '').strip()
            
            if not valve_serial and not actuator_serial:
                return JsonResponse({'iserror': True, 'message': 'At least one serial number is required'})
            
            # Collect all component status data
            cycle_data = {
                'valve_serial_number': valve_serial,
                'actuator_serial_number': actuator_serial,
                'air_shell_status': request.POST.get('air_shell_status', ''),
                'hydro_shell_status': request.POST.get('hydro_shell_status', ''),
                'air_seat_a_status': request.POST.get('air_seat_a_status', ''),
                'air_seat_b_status': request.POST.get('air_seat_b_status', ''),
                'valve_status': request.POST.get('valve_status', ''),
                'air_shell_result': request.POST.get('air_shell_result', ''),
                'hydro_shell_result': request.POST.get('hydro_shell_result', ''),
                'air_seat_a_result': request.POST.get('air_seat_a_result', ''),
                'air_seat_b_result': request.POST.get('air_seat_b_result', ''),
                'valve_result': request.POST.get('valve_result', ''),
                'air_shell_remarks': request.POST.get('air_shell_remarks', ''),
                'hydro_shell_remarks': request.POST.get('hydro_shell_remarks', ''),
                'air_seat_a_remarks': request.POST.get('air_seat_a_remarks', ''),
                'air_seat_b_remarks': request.POST.get('air_seat_b_remarks', ''),
                'valve_remarks': request.POST.get('valve_remarks', ''),
                'cycle_status': request.POST.get('cycle_status', ''),
                'completion_date': request.POST.get('completion_date', ''),
                'cycle_remarks': request.POST.get('cycle_remarks', ''),
                'created_at': timezone.now(),
                'updated_at': timezone.now()
            }
            
            with connection.cursor() as cursor:
                # Insert or update cycle status data
                cursor.execute("""
                    INSERT INTO valve_test_status (
                        valve_serial_number, actuator_serial_number, air_shell_status,
                        hydro_shell_status, air_seat_a_status, air_seat_b_status,
                        valve_status, air_shell_result, hydro_shell_result,
                        air_seat_a_result, air_seat_b_result, valve_result,
                        air_shell_remarks, hydro_shell_remarks, air_seat_a_remarks,
                        air_seat_b_remarks, valve_remarks, cycle_status,
                        completion_date, cycle_remarks, created_at, updated_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )
                """, [
                    cycle_data['valve_serial_number'], cycle_data['actuator_serial_number'],
                    cycle_data['air_shell_status'], cycle_data['hydro_shell_status'],
                    cycle_data['air_seat_a_status'], cycle_data['air_seat_b_status'],
                    cycle_data['valve_status'], cycle_data['air_shell_result'],
                    cycle_data['hydro_shell_result'], cycle_data['air_seat_a_result'],
                    cycle_data['air_seat_b_result'], cycle_data['valve_result'],
                    cycle_data['air_shell_remarks'], cycle_data['hydro_shell_remarks'],
                    cycle_data['air_seat_a_remarks'], cycle_data['air_seat_b_remarks'],
                    cycle_data['valve_remarks'], cycle_data['cycle_status'],
                    cycle_data['completion_date'], cycle_data['cycle_remarks'],
                    cycle_data['created_at'], cycle_data['updated_at']
                ])
                
                return JsonResponse({
                    'iserror': False, 
                    'message': 'Cycle status saved successfully',
                    'valve_status': cycle_data['valve_status'],
                    'cycle_status': cycle_data['cycle_status']
                })
                
        except Exception as e:
            logger.error(f"Error saving cycle status: {str(e)}")
            return JsonResponse({'iserror': True, 'message': 'Error saving cycle status data'})
    
    return JsonResponse({'iserror': True, 'message': 'Invalid request method'})

@csrf_exempt
def get_real_time_status(request):
    """Get real-time status for monitoring"""
    if request.method == 'POST':
        try:
            valve_serial = request.POST.get('valve_serial', '').strip()
            actuator_serial = request.POST.get('actuator_serial', '').strip()
            
            if not valve_serial and not actuator_serial:
                return JsonResponse({'success': False, 'message': 'Serial number required'})
            
            with connection.cursor() as cursor:
                if valve_serial and actuator_serial:
                    cursor.execute("""
                        SELECT air_shell_status, hydro_shell_status, air_seat_a_status, 
                               air_seat_b_status, valve_status, cycle_status
                        FROM valve_test_status 
                        WHERE valve_serial_number = %s AND actuator_serial_number = %s
                        ORDER BY created_at DESC LIMIT 1
                    """, [valve_serial, actuator_serial])
                elif valve_serial:
                    cursor.execute("""
                        SELECT air_shell_status, hydro_shell_status, air_seat_a_status, 
                               air_seat_b_status, valve_status, cycle_status
                        FROM valve_test_status 
                        WHERE valve_serial_number = %s
                        ORDER BY created_at DESC LIMIT 1
                    """, [valve_serial])
                else:
                    cursor.execute("""
                        SELECT air_shell_status, hydro_shell_status, air_seat_a_status, 
                               air_seat_b_status, valve_status, cycle_status
                        FROM valve_test_status 
                        WHERE actuator_serial_number = %s
                        ORDER BY created_at DESC LIMIT 1
                    """, [actuator_serial])
                
                result = cursor.fetchone()
                
                if result:
                    data = {
                        'air_shell_status': result[0],
                        'hydro_shell_status': result[1],
                        'air_seat_a_status': result[2],
                        'air_seat_b_status': result[3],
                        'valve_status': result[4],
                        'cycle_status': result[5]
                    }
                    return JsonResponse({'success': True, 'data': data})
                else:
                    return JsonResponse({'success': False, 'message': 'No status data found'})
                    
        except Exception as e:
            logger.error(f"Error getting real-time status: {str(e)}")
            return JsonResponse({'success': False, 'message': 'Error retrieving status data'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


